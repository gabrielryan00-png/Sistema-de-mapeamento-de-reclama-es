"""Agente autônomo de ouvidorias.

Usa Groq (llama-3.3-70b-versatile) com function calling para:
  1. Listar e-mails não processados
  2. Baixar e ler PDFs anexados
  3. Classificar como OUVIDORIA ou RESPOSTA com raciocínio contextual
  4. Salvar/vincular no Excel automaticamente
  5. Mover e-mails para as pastas Gmail corretas

Uso direto:
    python agente_ouvidoria.py

Uso via scheduler (dashboard_server.py):
    import agente_ouvidoria
    agente_ouvidoria.executar_ciclo(cfg, log_func)
"""

import os, json, re, time, shutil, tempfile
from datetime import date, datetime, timedelta
from typing import Optional

_BASE     = os.path.dirname(os.path.abspath(__file__))
_CFG_FILE = os.path.join(_BASE, 'config.json')

IMAP_HOST       = 'imap.gmail.com'
IMAP_PORT       = 993
LABEL_OUVIDORIAS = 'Ouvidorias/Ouvidorias'
LABEL_RESPOSTAS  = 'Ouvidorias/Respostas'
LABEL_PROCESSADO = 'Ouvidorias/Processado'
PRAZO_PADRAO     = 10

# ── Tool schemas (Groq function calling) ──────────────────────────────────────
_TOOLS = [
    {
        'type': 'function',
        'function': {
            'name': 'listar_emails_pendentes',
            'description': 'Lista e-mails não processados com PDFs na caixa de ouvidorias.',
            'parameters': {'type': 'object', 'properties': {}, 'required': []},
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'baixar_e_ler_pdf',
            'description': 'Baixa o PDF de um e-mail e retorna o texto extraído.',
            'parameters': {
                'type': 'object',
                'properties': {
                    'uid':      {'type': 'integer', 'description': 'UID do e-mail'},
                    'filename': {'type': 'string',  'description': 'Nome do arquivo PDF'},
                },
                'required': ['uid', 'filename'],
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'buscar_ouvidoria_por_protocolo',
            'description': 'Busca registro existente no Excel pelo número de protocolo.',
            'parameters': {
                'type': 'object',
                'properties': {
                    'protocolo': {'type': 'string', 'description': 'Número do protocolo (ex: 002799/2026)'},
                },
                'required': ['protocolo'],
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'salvar_ouvidoria',
            'description': (
                'Salva uma nova ouvidoria no Excel e move o PDF para a pasta correta. '
                'Use após classificar o documento como OUVIDORIA.'
            ),
            'parameters': {
                'type': 'object',
                'properties': {
                    'uid':              {'type': 'integer', 'description': 'UID do e-mail de origem'},
                    'pdf_filename':     {'type': 'string',  'description': 'Nome do PDF anexo'},
                    'protocolo':        {'type': 'string'},
                    'unidade':          {'type': 'string'},
                    'reclamante':       {'type': 'string'},
                    'data_recebimento': {'type': 'string', 'description': 'DD/MM/AAAA'},
                    'assunto':          {'type': 'string'},
                    'observacoes':      {'type': 'string'},
                },
                'required': ['uid', 'pdf_filename', 'protocolo'],
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'vincular_resposta',
            'description': (
                'Vincula um PDF de resposta a uma ouvidoria existente, '
                'atualiza status para RESPONDIDA. '
                'Use após classificar o documento como RESPOSTA.'
            ),
            'parameters': {
                'type': 'object',
                'properties': {
                    'uid':              {'type': 'integer', 'description': 'UID do e-mail de origem'},
                    'pdf_filename':     {'type': 'string',  'description': 'Nome do PDF anexo'},
                    'protocolo':        {'type': 'string'},
                    'data_respondida':  {'type': 'string', 'description': 'DD/MM/AAAA'},
                },
                'required': ['uid', 'pdf_filename', 'protocolo'],
            },
        },
    },
    {
        'type': 'function',
        'function': {
            'name': 'mover_e_marcar_processado',
            'description': 'Move o e-mail para a label Gmail correta e marca como processado.',
            'parameters': {
                'type': 'object',
                'properties': {
                    'uid':  {'type': 'integer'},
                    'tipo': {'type': 'string', 'enum': ['OUVIDORIA', 'RESPOSTA', 'IGNORAR']},
                },
                'required': ['uid', 'tipo'],
            },
        },
    },
]

_SYSTEM = """\
Você é um agente autônomo de gestão de ouvidorias municipais de saúde em Suzano/SP.

Sua missão: processar todos os e-mails pendentes da caixa de ouvidorias, classificar \
os PDFs e atualizar o sistema de registros.

FLUXO obrigatório para cada e-mail:
1. listar_emails_pendentes → veja os e-mails disponíveis.
2. baixar_e_ler_pdf → leia o texto do PDF de cada e-mail.
3. Analise o texto e decida: OUVIDORIA ou RESPOSTA?
   - OUVIDORIA: vem de cidadão/Controladoria; contém "EMENTA DA DEMANDA", "Lei 13.460", P.O.
   - RESPOSTA: vem da unidade de saúde; contém assinatura de gerente/enfermeira.
4a. Se OUVIDORIA → salvar_ouvidoria com todos os dados extraídos.
4b. Se RESPOSTA  → buscar_ouvidoria_por_protocolo para encontrar o registro existente,
                   depois vincular_resposta.
5. mover_e_marcar_processado para finalizar cada e-mail.
6. Repita para todos os e-mails. Quando não houver mais, encerre.

REGRAS:
- Extraia protocolo no formato NNN.../AAAA. Se não encontrar, use "NÃO IDENTIFICADO".
- Extraia unidade (nome da USF) do texto do PDF.
- Se for RESPOSTA e não encontrar o protocolo no sistema, use vincular_resposta \
  com protocolo "NÃO IDENTIFICADO".
- Nunca invente dados — use string vazia se não encontrar."""


# ── Agent class ────────────────────────────────────────────────────────────────
class AgenteOuvidoria:
    def __init__(self, cfg: dict, log=print, reprocessar=False):
        self.cfg         = cfg
        self.log         = log
        self.reprocessar = reprocessar   # True = busca todos os e-mails, não só não lidos
        self.imap        = None
        self._cache: dict = {}   # (uid, filename) → local_path
        self._tmpdir = tempfile.mkdtemp(prefix='agente_ouv_')
        self._ids_proc: set = set()

    def __enter__(self):
        self._connect()
        return self

    def __exit__(self, *_):
        self._close()

    # ── IMAP ──────────────────────────────────────────────────────────────────
    def _connect(self):
        import imapclient
        self.imap = imapclient.IMAPClient(IMAP_HOST, port=IMAP_PORT, ssl=True)
        self.imap.login(self.cfg['email'], self.cfg['senha_app'])
        self.log('📬 Conectado ao Gmail')
        for pasta in (LABEL_OUVIDORIAS, LABEL_RESPOSTAS, LABEL_PROCESSADO):
            self._garantir_pasta(pasta)

    def _close(self):
        try:
            if self.imap:
                self.imap.logout()
        except Exception:
            pass
        try:
            shutil.rmtree(self._tmpdir, ignore_errors=True)
        except Exception:
            pass

    def _garantir_pasta(self, nome: str):
        try:
            pastas = [f[2] for f in self.imap.list_folders()]
            if nome not in pastas:
                self.imap.create_folder(nome)
        except Exception:
            pass

    # ── Groq call ─────────────────────────────────────────────────────────────
    def _groq(self, messages: list) -> dict:
        import requests
        for tentativa in range(3):
            resp = requests.post(
                'https://api.groq.com/openai/v1/chat/completions',
                headers={'Authorization': f'Bearer {self.cfg["groq_api_key"]}'},
                json={
                    'model':       'llama-3.3-70b-versatile',
                    'messages':    messages,
                    'tools':       _TOOLS,
                    'tool_choice': 'auto',
                    'temperature': 0.1,
                    'max_tokens':  1024,
                },
                timeout=40,
            )
            if resp.status_code == 429:
                wait = int(resp.headers.get('retry-after', 15))
                self.log(f'  ⏳ Rate limit — aguardando {wait}s')
                time.sleep(wait)
                continue
            resp.raise_for_status()
            return resp.json()
        raise RuntimeError('Groq indisponível após 3 tentativas')

    # ── Agent loop ────────────────────────────────────────────────────────────
    def executar(self):
        modo = 'REPROCESSAR TODOS os e-mails' if self.reprocessar else 'processar e-mails pendentes'
        messages = [
            {'role': 'system', 'content': _SYSTEM},
            {'role': 'user',   'content': f'Processe: {modo} da caixa de ouvidorias.'},
        ]
        self.log(f'🤖 Agente iniciado — modo: {modo}')

        for iteracao in range(40):
            messages = self._trim_history(messages)
            response = self._groq(messages)
            choice   = response['choices'][0]
            msg      = choice['message']
            messages.append(msg)

            tool_calls = msg.get('tool_calls') or []

            if not tool_calls:
                self.log(f'🤖 Agente: {msg.get("content", "Concluído.")}')
                break

            for tc in tool_calls:
                name = tc['function']['name']
                try:
                    args = json.loads(tc['function']['arguments'] or '{}') or {}
                except (json.JSONDecodeError, TypeError):
                    args = {}

                self.log(f'  🔧 {name}({", ".join(f"{k}={v!r}" for k, v in args.items())})')

                result = self._exec(name, args)
                preview = json.dumps(result, ensure_ascii=False)[:300]
                self.log(f'     → {preview}')

                messages.append({
                    'role':         'tool',
                    'tool_call_id': tc['id'],
                    'content':      json.dumps(result, ensure_ascii=False),
                })
        else:
            self.log('⚠️  Limite de iterações atingido')

    def _trim_history(self, messages: list) -> list:
        """Remove pares antigos de tool_call/tool quando o histórico passa de ~40k chars."""
        total = sum(len(json.dumps(m)) for m in messages)
        if total < 40_000:
            return messages

        # Mantém system + user iniciais e os últimos 12 mensagens
        head = [m for m in messages if m['role'] in ('system', 'user')][:2]
        tail = messages[-12:]
        # Garante que não começa com role=tool (precisa do assistant antes)
        while tail and tail[0].get('role') == 'tool':
            tail = tail[1:]
        trimmed = head + tail
        self.log(f'  🗜 Histórico compactado: {len(messages)} → {len(trimmed)} mensagens')
        return trimmed

    def _exec(self, name: str, args: dict) -> dict:
        fn = getattr(self, f'_t_{name}', None)
        if fn is None:
            return {'error': f'ferramenta desconhecida: {name}'}
        try:
            return fn(**args)
        except Exception as e:
            return {'error': str(e)}

    # ── Tools ─────────────────────────────────────────────────────────────────
    def _t_listar_emails_pendentes(self) -> dict:
        import pyzmail
        from ouvidoriagmail import obter_message_id

        # Carrega IDs já processados (para deduplicação no modo normal)
        try:
            self.imap.select_folder(LABEL_PROCESSADO)
            uids_proc = self.imap.search(['ALL'])
            if uids_proc:
                fetched = self.imap.fetch(uids_proc, ['BODY.PEEK[HEADER.FIELDS (MESSAGE-ID)]'])
                if fetched:
                    for uid, data in fetched.items():
                        raw = data.get(b'BODY[HEADER.FIELDS (MESSAGE-ID)]', b'')
                        m = re.search(rb'Message-ID:\s*(<[^>]+>)', raw, re.IGNORECASE)
                        if m:
                            self._ids_proc.add(m.group(1).decode().strip())
        except Exception:
            pass

        # Pastas a varrer
        pastas = ['INBOX']
        if self.reprocessar:
            pastas += [LABEL_OUVIDORIAS, LABEL_RESPOSTAS, LABEL_PROCESSADO]

        rem      = self.cfg.get('remetente_ouvidoria', '').strip()
        vistos   = set()
        lista    = []

        for pasta in pastas:
            try:
                self.imap.select_folder(pasta)
            except Exception:
                continue

            criteria = [] if self.reprocessar else ['UNSEEN']
            if rem:
                criteria = ['FROM', rem] + criteria
            if not criteria:
                criteria = ['ALL']

            uids = self.imap.search(criteria)

            for uid in uids:
                if uid in vistos:
                    continue
                vistos.add(uid)

                try:
                    fetched = self.imap.fetch([uid], ['BODY.PEEK[]'])
                    if not fetched or uid not in fetched:
                        continue
                    raw = fetched[uid].get(b'BODY[]') or fetched[uid].get(b'BODY.PEEK[]')
                    if not raw:
                        continue
                    msg = pyzmail.PyzMessage.factory(raw)
                    mid = obter_message_id(msg)
                except Exception:
                    continue

                if not self.reprocessar and mid and mid in self._ids_proc:
                    continue

                pdfs = [p.filename for p in msg.mailparts
                        if p.filename and p.filename.lower().endswith('.pdf')]
                if not pdfs:
                    continue

                # Filtra: e-mail deve ter "ouvidoria" no assunto, no corpo ou no nome do PDF
                _TERMOS = ('ouvidoria', 'ouv.', 'p.o.', 'p.o ', 'po ')
                assunto_lower = (msg.get_subject() or '').lower()
                pdfs_lower    = ' '.join(pdfs).lower()
                corpo = ''
                if msg.text_part:
                    try:
                        corpo = (msg.text_part.get_payload() or b'').decode(
                            msg.text_part.charset or 'utf-8', errors='ignore').lower()
                    except Exception:
                        pass
                tem_termo = any(t in assunto_lower or t in pdfs_lower or t in corpo
                                for t in _TERMOS)
                if not tem_termo:
                    continue

                lista.append({
                    'uid':        int(uid),
                    'pasta':      pasta,
                    'assunto':    msg.get_subject() or '',
                    'data':       str(msg.get_decoded_header('Date', ''))[:30],
                    'pdfs':       pdfs,
                    'message_id': mid,
                })

        return {'emails': lista, 'total': len(lista)}

    def _t_baixar_e_ler_pdf(self, uid: int, filename: str) -> dict:
        import pyzmail
        from ouvidoriagmail import ler_pdf as _ler

        pastas = ['INBOX', LABEL_OUVIDORIAS, LABEL_RESPOSTAS, LABEL_PROCESSADO]
        for pasta in pastas:
            try:
                self.imap.select_folder(pasta)
                fetched = self.imap.fetch([uid], ['BODY[]'])
                if not fetched or uid not in fetched:
                    continue
                raw = fetched[uid].get(b'BODY[]')
                if not raw:
                    continue
            except Exception:
                continue

            msg = pyzmail.PyzMessage.factory(raw)
            for part in msg.mailparts:
                if part.filename == filename:
                    local = os.path.join(self._tmpdir, f'{uid}_{filename}')
                    with open(local, 'wb') as f:
                        f.write(part.get_payload())
                    self._cache[(uid, filename)] = local
                    texto = _ler(local)
                    return {'texto': texto[:2000], 'chars': len(texto)}

        return {'error': f'PDF {filename!r} não encontrado no e-mail {uid}'}

    def _t_buscar_ouvidoria_por_protocolo(self, protocolo: str) -> dict:
        from openpyxl import load_workbook
        path = self._excel_path()
        if not os.path.exists(path):
            return {'encontrado': False}

        wb = load_workbook(path, read_only=True)
        if 'Ouvidorias' not in wb.sheetnames:
            wb.close()
            return {'encontrado': False}

        ws      = wb['Ouvidorias']
        headers = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]
        digits  = re.sub(r'\D', '', protocolo)

        for row in ws.iter_rows(min_row=2, values_only=True):
            cell = str(row[0] or '').strip()
            if digits and digits in re.sub(r'\D', '', cell):
                rec = {h: str(v or '') for h, v in zip(headers, row) if h}
                wb.close()
                return {'encontrado': True, 'registro': rec}

        wb.close()
        return {'encontrado': False}

    def _t_salvar_ouvidoria(self, uid: int, pdf_filename: str, protocolo: str,
                             unidade: str = '', reclamante: str = '',
                             data_recebimento: str = '', assunto: str = '',
                             observacoes: str = '') -> dict:
        local = self._cache.get((uid, pdf_filename))
        arquivo_final = ''

        if local and os.path.exists(local):
            pasta_base = self.cfg.get('pasta_base', 'ouvidorias')
            destino    = os.path.join(_BASE, pasta_base, 'ouvidorias', os.path.basename(local))
            os.makedirs(os.path.dirname(destino), exist_ok=True)
            shutil.move(local, destino)
            arquivo_final = os.path.basename(destino)
            self._cache[(uid, pdf_filename)] = destino

        prazo = self._calcular_prazo(data_recebimento)

        self._append_excel({
            'Protocolo':        protocolo or 'NÃO IDENTIFICADO',
            'Unidade':          unidade,
            'Reclamante':       reclamante,
            'Data Recebimento': data_recebimento,
            'Prazo Resposta':   prazo,
            'Assunto':          assunto,
            'Arquivo':          arquivo_final,
            'Status':           'PENDENTE',
            'Observações':      observacoes,
        })
        return {'ok': True, 'protocolo': protocolo, 'arquivo': arquivo_final}

    def _t_vincular_resposta(self, uid: int, pdf_filename: str, protocolo: str,
                              data_respondida: str = '') -> dict:
        local = self._cache.get((uid, pdf_filename))
        arquivo_final = ''

        if local and os.path.exists(local):
            pasta_base = self.cfg.get('pasta_base', 'ouvidorias')
            destino    = os.path.join(_BASE, pasta_base, 'respostas', os.path.basename(local))
            os.makedirs(os.path.dirname(destino), exist_ok=True)
            shutil.move(local, destino)
            arquivo_final = os.path.basename(destino)
            self._cache[(uid, pdf_filename)] = destino

        data_resp = data_respondida or datetime.now().strftime('%d/%m/%Y')
        ok = self._patch_excel(protocolo, {
            'Status':           'RESPONDIDA',
            'Arquivo Resposta': arquivo_final,
            'Data Respondida':  data_resp,
        })

        if not ok:
            # Protocolo não encontrado — salva linha nova marcada como resposta
            self._append_excel({
                'Protocolo':        protocolo or 'NÃO IDENTIFICADO',
                'Status':           'RESPONDIDA',
                'Arquivo Resposta': arquivo_final,
                'Data Respondida':  data_resp,
                'Observações':      'Resposta sem ouvidoria correspondente no sistema',
            })

        return {'ok': True, 'vinculado': ok, 'arquivo': arquivo_final}

    def _t_mover_e_marcar_processado(self, uid: int, tipo: str) -> dict:
        # Ignora UIDs inválidos/fabricados pelo modelo
        if uid <= 0 or uid > 9_999_999:
            return {'ok': False, 'error': f'UID inválido: {uid}'}

        try:
            pastas = ['INBOX', LABEL_OUVIDORIAS, LABEL_RESPOSTAS, LABEL_PROCESSADO]
            encontrado = False
            for pasta in pastas:
                try:
                    self.imap.select_folder(pasta)
                    fetched = self.imap.fetch([uid], ['BODY.PEEK[HEADER.FIELDS (FROM)]'])
                    if fetched and uid in fetched:
                        encontrado = True
                        break
                except Exception:
                    continue

            if not encontrado:
                return {'ok': False, 'error': f'E-mail UID {uid} não encontrado'}

            if tipo == 'OUVIDORIA':
                self.imap.add_gmail_labels([uid], [LABEL_OUVIDORIAS])
            elif tipo == 'RESPOSTA':
                self.imap.add_gmail_labels([uid], [LABEL_RESPOSTAS])

            if tipo != 'IGNORAR':
                self.imap.add_gmail_labels([uid], [LABEL_PROCESSADO])

        except Exception as e:
            return {'ok': False, 'error': str(e)}

        return {'ok': True, 'uid': uid, 'tipo': tipo}

    # ── Excel helpers ─────────────────────────────────────────────────────────
    def _excel_path(self) -> str:
        pasta = self.cfg.get('pasta_base', 'ouvidorias')
        return os.path.join(_BASE, pasta, 'ouvidorias.xlsx')

    def _calcular_prazo(self, data_rec: str) -> str:
        try:
            d  = datetime.strptime(data_rec, '%d/%m/%Y').date()
            prazo_dias = int(self.cfg.get('prazo_dias', PRAZO_PADRAO))
            return (d + timedelta(days=prazo_dias)).strftime('%d/%m/%Y')
        except Exception:
            return ''

    def _append_excel(self, data: dict):
        from openpyxl import load_workbook, Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from constantes import COLUNAS_OUV

        path = self._excel_path()
        os.makedirs(os.path.dirname(path), exist_ok=True)

        wb = load_workbook(path) if os.path.exists(path) else Workbook()
        if not os.path.exists(path):
            wb.remove(wb.active)

        if 'Ouvidorias' not in wb.sheetnames:
            ws = wb.create_sheet('Ouvidorias')
            for i, col in enumerate(COLUNAS_OUV + ['Reclamante', 'Canal'], 1):
                c = ws.cell(1, i, col)
                c.font      = Font(bold=True, color='FFFFFF', name='Arial', size=11)
                c.fill      = PatternFill('solid', start_color='2B5597')
                c.alignment = Alignment(horizontal='center', vertical='center')
            ws.freeze_panes = 'A2'
        else:
            ws = wb['Ouvidorias']

        headers = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]
        row_n   = ws.max_row + 1
        for col, val in data.items():
            if col in headers:
                ws.cell(row_n, headers.index(col) + 1, str(val) if val else '')

        wb.save(path)
        wb.close()

    def _patch_excel(self, protocolo: str, updates: dict) -> bool:
        from openpyxl import load_workbook
        path = self._excel_path()
        if not os.path.exists(path):
            return False

        wb = load_workbook(path)
        if 'Ouvidorias' not in wb.sheetnames:
            wb.close()
            return False

        ws      = wb['Ouvidorias']
        headers = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]
        digits  = re.sub(r'\D', '', protocolo)

        for row in range(2, ws.max_row + 1):
            cell = str(ws.cell(row, 1).value or '').strip()
            if digits and digits in re.sub(r'\D', '', cell):
                for col, val in updates.items():
                    if col in headers:
                        ws.cell(row, headers.index(col) + 1, str(val) if val else '')
                wb.save(path)
                wb.close()
                return True

        wb.close()
        return False


# ── Public API ─────────────────────────────────────────────────────────────────
def executar_ciclo(cfg: Optional[dict] = None, log=print,
                   reprocessar: bool = False) -> dict:
    """Entry point para uso pelo scheduler ou linha de comando."""
    if cfg is None:
        try:
            with open(_CFG_FILE, encoding='utf-8') as f:
                cfg = json.load(f)
        except Exception as e:
            log(f'❌ Erro ao ler config.json: {e}')
            return {'ok': False, 'error': str(e)}

    if not cfg.get('groq_api_key', '').strip():
        log('❌ groq_api_key não configurado')
        return {'ok': False, 'error': 'groq_api_key ausente'}

    try:
        with AgenteOuvidoria(cfg, log, reprocessar=reprocessar) as agente:
            agente.executar()
        return {'ok': True}
    except Exception as e:
        log(f'❌ Agente encerrou com erro: {e}')
        return {'ok': False, 'error': str(e)}


if __name__ == '__main__':
    import sys
    reprocessar = '--reprocessar' in sys.argv

    def _log(msg):
        print(msg, flush=True)

    resultado = executar_ciclo(log=_log, reprocessar=reprocessar)
    sys.exit(0 if resultado['ok'] else 1)
