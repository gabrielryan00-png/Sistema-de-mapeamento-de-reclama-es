import os
import re
import difflib
import unicodedata
import smtplib
from datetime import datetime, date, timedelta
from email.message import EmailMessage
from email.utils import formatdate
from typing import Optional, Dict, List

import pdfplumber
import pytesseract
from pdf2image import convert_from_path
import pandas as pd
import imapclient
import pyzmail
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =========================
# CONFIG GMAIL
# =========================
EMAIL     = ""   # Configurar em config.json → "email"
SENHA_APP = ""   # Configurar em config.json → "senha_app"
# Para gerar a senha de app: myaccount.google.com → Segurança → Senhas de app
# (requer verificação em duas etapas ativada)

REMETENTE_OUVIDORIA = ""  # e-mail de quem envia as ouvidorias (deixe "" para qualquer)

IMAP_HOST = "imap.gmail.com"
IMAP_PORT = 993
SMTP_HOST = "smtp.gmail.com"
SMTP_PORT = 465  # SSL

# Labels Gmail
LABEL_OUVIDORIAS = "Ouvidorias/Ouvidorias"
LABEL_RESPOSTAS  = "Ouvidorias/Respostas"
LABEL_PROCESSADO = "Ouvidorias/Processado"

# Pastas locais
PASTA_BASE        = "ouvidorias"
PASTA_OUVIDORIAS  = os.path.join(PASTA_BASE, "ouvidorias")
PASTA_RESPOSTAS   = os.path.join(PASTA_BASE, "respostas")

# Aliases para compatibilidade com restante do código
IMAP_OUVIDORIAS = LABEL_OUVIDORIAS
IMAP_RESPOSTAS  = LABEL_RESPOSTAS
IMAP_PROCESSADO = LABEL_PROCESSADO

RELATORIO = os.path.join(PASTA_BASE, "ouvidorias.xlsx")
PRAZO_DIAS = 10

os.makedirs(PASTA_OUVIDORIAS, exist_ok=True)
os.makedirs(PASTA_RESPOSTAS, exist_ok=True)

# Interactive prompts moved to the script entrypoint to make this module import-safe.

# =========================
# PDF: EXTRAÇÃO + OCR
# =========================
def ler_pdf(pdf_path: str) -> str:
    texto = ""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for p in pdf.pages:
                t = p.extract_text()
                if t:
                    texto += t + "\n"
    except Exception as e:
        print(f"  Erro pdfplumber: {e}")

    if len(texto.strip()) < 100:
        print("  → OCR...")
        try:
            imagens = convert_from_path(pdf_path)
            ocr = ""
            for img in imagens:
                ocr += pytesseract.image_to_string(img, lang="por") + "\n"
            if len(ocr.strip()) > len(texto.strip()):
                texto = ocr
        except Exception as e:
            print(f"  Erro OCR: {e}")

    return texto.strip()

# =========================
# EXTRAÇÃO DE DADOS DA OUVIDORIA
# =========================
def extrair_protocolo(texto: str) -> str:
    """
    Extrai o número de protocolo. Suporta os formatos encontrados nos documentos:
      - "P.O. 003279/2026"  → "P.O. 003279/2026"  (ouvidoria da Controladoria)
      - "Ref. PO 2278"      → "P.O. 2278"          (resposta da unidade)
      - "P.O. 003279/2026 - OUVIDORIA" (linha de cabeçalho)
    """
    # Prioridade 1: P.O. XXXXXX/AAAA — formato canônico da Controladoria
    m = re.search(r"P\.?\s*O\.?\s*[:\-]?\s*(\d{3,})\s*[/\-]\s*(\d{4})", texto, re.IGNORECASE)
    if m:
        return f"P.O. {m.group(1)}/{m.group(2)}"

    # Prioridade 2: "Ref. PO XXXX" ou "Ref: P.O. XXXX" — cabeçalho de resposta
    m = re.search(r"Ref\.?\s*[:\-]?\s*P\.?\s*O\.?\s*[\.\-\s]*(\d{3,})", texto, re.IGNORECASE)
    if m:
        return f"P.O. {m.group(1)}"

    # Prioridade 3: "PO XXXX" isolado (sem Ref., sem ano)
    m = re.search(r"\bP\.?\s*O\.?\s+(\d{3,})\b", texto, re.IGNORECASE)
    if m:
        return f"P.O. {m.group(1)}"

    # Prioridade 4: campo genérico "Protocolo: XXXX"
    m = re.search(r"(?:Protocolo|Proc\.?|N[uú]mero)[:\s#]*(\d{4,}[\s/.-]\d{2,4})", texto, re.IGNORECASE)
    if m:
        return m.group(1).strip()

    return ""

# =========================
# CATÁLOGO DE USFs — IDENTIFICAÇÃO POR APROXIMAÇÃO
# =========================
# Para adicionar novas unidades: inclua um novo dict com "nome" e "aliases".
# Aliases devem conter variações do nome, apelidos e nomes de bairro.
CATALOGO_USF = [
    {
        "nome": "USF Vereador Marsal Lopes Rosa",
        "aliases": [
            "vereador marsal lopes rosa", "marsal lopes rosa", "marsal rosa",
            "vila amorim", "usf vila amorim", "ubs vila amorim",
        ],
    },
    {
        "nome": "USF Dr. Eduardo Nakamura",
        "aliases": [
            "eduardo nakamura", "dr eduardo nakamura", "doutor eduardo nakamura",
            "nakamura", "dr nakamura", "jardim nakamura",
        ],
    },
    {
        "nome": "USF Jardim Europa",
        "aliases": [
            "jardim europa", "jd europa", "usf europa", "ubs europa",
        ],
    },
    {
        "nome": "USF Jardim do Lago",
        "aliases": [
            "jardim do lago", "jd do lago", "jardim lago", "usf lago", "ubs lago",
        ],
    },
    {
        "nome": "USF Manuel Evangelista de Oliveira",
        "aliases": [
            "manuel evangelista de oliveira", "manuel evangelista", "manoel evangelista",
            "jardim sao jose", "jardim são jose", "jardim são josé",
            "jardim s jose", "jd sao jose",
        ],
    },
    {
        "nome": "USF Onesia Benedita Miguel",
        "aliases": [
            "onesia benedita miguel", "onesia benedita", "onesia miguel",
            "jardim suzanopolis", "jardim suzanópolis", "suzanopolis", "suzanópolis",
        ],
    },
    {
        "nome": "USF Maria Jose Lima Souza",
        "aliases": [
            "maria jose lima souza", "maria jose lima", "maria jose souza",
            "jardim ikeda", "ikeda",
        ],
    },
    {
        "nome": "USF Antonio Marques de Carvalho",
        "aliases": [
            "antonio marques de carvalho", "antonio marques", "antônio marques",
            "jardim maite", "maite",
        ],
    },
    {
        "nome": "USF Marcelino Maria Rodrigues",
        "aliases": [
            "marcelino maria rodrigues", "marcelino maria", "marcelino rodrigues",
            "jardim brasil", "jd brasil",
        ],
    },
    {
        "nome": "USF Vereador Gregório Bonifácio da Silva",
        "aliases": [
            "vereador gregorio bonifacio da silva", "gregorio bonifacio da silva",
            "gregorio bonifacio", "gregório bonifácio", "vereador gregório",
            "vila fatima", "vila fátima",
        ],
    },
    {
        "nome": "USF Recanto São José",
        "aliases": [
            "recanto sao jose", "recanto são jose", "recanto são josé",
            "recanto s jose", "usf recanto", "ubs recanto", "recanto",
        ],
    },
    {
        "nome": "USF Jardim Revista",
        "aliases": [
            "jardim revista", "jd revista", "usf revista", "ubs revista",
        ],
    },
]

def _normalizar_texto(texto: str) -> str:
    """Remove acentos, pontuação e normaliza espaços para comparação."""
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = re.sub(r"[^\w\s]", " ", texto)
    texto = re.sub(r"\s+", " ", texto).strip().lower()
    return texto

def identificar_usf(texto: str, threshold: float = 0.72) -> Optional[str]:
    """
    Identifica a USF mencionada no texto usando três estratégias em cascata:
      1) Correspondência exata por substring normalizada
      2) Fuzzy matching por janelas deslizantes (tolera erros de digitação e grafia)
    Retorna o nome canônico da USF ou None se não identificada.
    """
    texto_norm = _normalizar_texto(texto)
    palavras   = texto_norm.split()

    melhor_nome  = None
    melhor_score = 0.0

    for usf in CATALOGO_USF:
        for alias in usf["aliases"]:
            alias_norm = _normalizar_texto(alias)

            
            if alias_norm in texto_norm:
                return usf["nome"]

            
            alias_palavras = alias_norm.split()
            n = len(alias_palavras)
            for i in range(max(1, len(palavras) - n + 1)):
                janela = " ".join(palavras[i:i+n])
                score  = difflib.SequenceMatcher(None, alias_norm, janela).ratio()
                if score > melhor_score:
                    melhor_score = score
                    melhor_nome  = usf["nome"]

    return melhor_nome if melhor_score >= threshold else None

def extrair_unidade(texto: str) -> str:
    """
    Identifica a unidade de saúde (USF/UBS) mencionada no texto.
    Usa o catálogo CATALOGO_USF com matching fuzzy para tolerar
    erros de grafia, ausência de acentos e uso de apelidos de bairro.
    """
    # Tentativa 1: catálogo com fuzzy matching (cobre a maior parte dos casos)
    usf = identificar_usf(texto)
    if usf:
        return usf

    # Tentativa 2: campo PARA: do cabeçalho (documentos da Controladoria)
    m = re.search(r"PARA\s*:\s*([^\n]{3,80})", texto, re.IGNORECASE)
    if m:
        dest = m.group(1).strip()
        if dest and dest.upper() not in ("", "OUVIDORIA", "CONTROLADORIA"):
            # Tenta passar o campo PARA pelo catálogo também
            usf_dest = identificar_usf(dest)
            if usf_dest:
                return usf_dest
            return dest

    return "NÃO IDENTIFICADA"

_MESES_PT = {
    'janeiro': 1, 'fevereiro': 2, 'marco': 3, 'março': 3,
    'abril': 4, 'maio': 5, 'junho': 6, 'julho': 7,
    'agosto': 8, 'setembro': 9, 'outubro': 10,
    'novembro': 11, 'dezembro': 12,
}

def extrair_data_documento(texto: str) -> Optional[date]:
    """
    Extrai a data do documento. Suporta:
      - "23-04-2026" / "23/04/2026"        (Controladoria)
      - "08 de abril de 2026"              (resposta de unidade)
      - "Suzano, 23-04-2026"
    """
    # Prioridade 1: dd-mm-aaaa ou dd/mm/aaaa
    for m in re.finditer(r"(\d{2})[-/](\d{2})[-/](\d{4})", texto):
        try:
            return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            continue

    # Prioridade 2: "8 de abril de 2026" — formato por extenso
    m = re.search(r"(\d{1,2})\s+de\s+(\w+)\s+de\s+(\d{4})", texto, re.IGNORECASE)
    if m:
        mes = _MESES_PT.get(m.group(2).lower().strip())
        if mes:
            try:
                return date(int(m.group(3)), mes, int(m.group(1)))
            except ValueError:
                pass

    # Prioridade 3: aaaa-mm-dd
    for m in re.finditer(r"(\d{4})[-/](\d{2})[-/](\d{2})", texto):
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            continue

    return None

def extrair_tipo(texto: str) -> str:
    """
    Classifica o PDF como 'OUVIDORIA' ou 'RESPOSTA'.

    Estratégia em duas camadas:
      1. Regras determinísticas de alta certeza (estrutura do documento)
      2. Pontuação por sinais configuráveis em constantes.py (fallback)
    """
    from constantes import SINAIS_RESPOSTA, SINAIS_OUVIDORIA, CLASSIFIER_WEIGHTS

    inicio = texto[:600]  # cabeçalho onde a identidade do documento fica mais clara

    # ── Regras determinísticas ─────────────────────────────────────────────────
    # Respostas de unidade sempre começam com "Resposta à Manifestação de Ouvidoria"
    if re.search(r"Resposta\s+à\s+Manifesta[çc][aã]o\s+de\s+Ouvidoria", inicio, re.IGNORECASE):
        return "RESPOSTA"

    # Cabeçalho típico da Controladoria: "EMENTA DA DEMANDA" ou "E M E N T A"
    if re.search(r"E\s*M\s*E\s*N\s*T\s*A\s+D\s*A\s+D\s*E\s*M\s*A\s*N\s*D\s*A", texto, re.IGNORECASE):
        return "OUVIDORIA"

    # Linha de encaminhamento da Controladoria
    if re.search(r"CONTROLADORIA\s+GERAL\s+DO\s+MUNIC", inicio, re.IGNORECASE) and \
       re.search(r"ENCAMINHAMENTO", texto, re.IGNORECASE):
        return "OUVIDORIA"

    # "Atenciosamente" + assinatura de gerente/enfermeira → resposta de unidade
    if re.search(r"Atenciosamente", texto, re.IGNORECASE) and \
       re.search(r"Enfermeira|Gerente|Coordenadora|Diretor", texto, re.IGNORECASE):
        return "RESPOSTA"

    # ── Pontuação por sinais (fallback) ────────────────────────────────────────
    def count_matches(patterns, text):
        s = 0
        for p in patterns:
            try:
                s += len(re.findall(p, text, flags=re.IGNORECASE))
            except re.error:
                s += text.upper().count(p.upper())
        return s

    score_resposta  = count_matches(SINAIS_RESPOSTA, texto)
    score_ouvidoria = count_matches(SINAIS_OUVIDORIA, texto)

    # Boost se início tem sinais fortes de resposta
    if any(x.upper() in inicio.upper() for x in
           ["EM ATENÇÃO", "EM RESPOSTA", "EM ATENCAO", "RETORNO À MANIFESTAÇÃO"]):
        score_resposta += CLASSIFIER_WEIGHTS.get("inicio_boost", 2)

    score_resposta  *= CLASSIFIER_WEIGHTS.get("resposta_weight", 1)
    score_ouvidoria *= CLASSIFIER_WEIGHTS.get("ouvidoria_weight", 1)

    if score_resposta > score_ouvidoria:
        return "RESPOSTA"
    if score_ouvidoria > score_resposta:
        return "OUVIDORIA"

    # Empate: "P.O. XXXXX" é exclusivo de ouvidoria
    if re.search(r"\bP\.?\s*O\.\s+\d{3,}", texto, re.IGNORECASE):
        return "OUVIDORIA"

    return "OUVIDORIA"

def extrair_assunto(texto: str) -> str:
    """
    Extrai o assunto/ementa da demanda.

    Estrutura do documento Controladoria:
      P.O. XXXXXX/AAAA - OUVIDORIA
      ATENDIMENTO INSATISFATÓRIO / SERVIÇO IRREGULAR   ← assunto destacado
      - E M E N T A   D A   D E M A N D A:
    """
    # Prioridade 1: linha destacada entre cabeçalho P.O. e EMENTA (estrutura da Controladoria)
    m = re.search(
        r"OUVIDORIA\s*\n\s*([A-ZÁÀÃÂÉÊÍÓÔÕÚ][^\n]{5,120})\s*\n",
        texto, re.IGNORECASE
    )
    if m:
        candidato = m.group(1).strip()
        # Exclui linhas que são claramente partes do cabeçalho
        if not re.search(r"PREFEITURA|CONTROLADORIA|ESTADO DE", candidato, re.IGNORECASE):
            return candidato[:120]

    # Prioridade 2: padrão "EMENTA DA DEMANDA:" seguido do assunto na mesma ou próxima linha
    m = re.search(
        r"E\s*M\s*E\s*N\s*T\s*A[^\n]*\n\s*([A-ZÁÀÃÂÉÊÍÓÔÕÚ][^\n]{5,120})",
        texto, re.IGNORECASE
    )
    if m:
        return m.group(1).strip()[:120]

    # Prioridade 3: categorias conhecidas diretamente no texto
    m = re.search(
        r"(ATENDIMENTO\s+INSATISFAT[^\n]{0,60}|FALTA\s*/\s*DEMORA[^\n]{0,80}"
        r"|ACESSO\s+A\s+[^\n]{0,60}|SERVI[CÇ]O\s+IRREGULAR[^\n]{0,60}"
        r"|RECLAMAÇÃO[^\n]{0,60}|DENÚNCIA[^\n]{0,60})",
        texto, re.IGNORECASE
    )
    if m:
        return m.group(1).strip()[:120]

    # Prioridade 4: campo genérico
    m = re.search(r"(?:Trata-se de|Objeto[:\s]+|Assunto[:\s]+)([^\n]{10,150})", texto, re.IGNORECASE)
    if m:
        return m.group(1).strip()[:120]

    return ""


def extrair_reclamante(texto: str, tipo: str = "OUVIDORIA") -> str:
    """
    Extrai o nome do reclamante (ou destinatário em respostas).

    Ouvidoria: "Me chamo Silene" / "Meu nome é X"
    Resposta:  "Prezada Sra. Elda Cristina de Lima"
    """
    if tipo == "RESPOSTA":
        # "Prezada Sra. Elda Cristina de Lima,"
        m = re.search(
            r"Prezad[ao]\s+Sr[ao]?\.?\s+([A-ZÁÀÃÂÉÊÍÓÔÕÚ][a-záàãâéêíóôõú]+(?:\s+[A-ZÁÀÃÂÉÊÍÓÔÕÚ][a-záàãâéêíóôõú]+){1,5})",
            texto
        )
        if m:
            return m.group(1).strip()

    # "Me chamo Silene tenho 32 anos"  → captura o primeiro nome/nomes
    m = re.search(
        r"[Mm]e\s+chamo\s+([A-ZÁÀÃÂÉÊÍÓÔÕÚ][a-záàãâéêíóôõú]+(?:\s+[A-ZÁÀÃÂÉÊÍÓÔÕÚ][a-záàãâéêíóôõú]+)*)",
        texto
    )
    if m:
        # Para "Me chamo Silene tenho 32 anos" → pega só o nome antes de "tenho"/"com"/"estou"
        nome = re.split(r'\s+(?:tenho|com|estou|tenho|sou|e\s+)\b', m.group(1), flags=re.IGNORECASE)[0]
        return nome.strip()

    # "Meu nome é / Meu nome:"
    m = re.search(
        r"[Mm]eu\s+nome\s+[eé][:\s]+([A-ZÁÀÃÂÉÊÍÓÔÕÚ][a-záàãâéêíóôõú]+(?:\s+[A-ZÁÀÃÂÉÊÍÓÔÕÚ][a-záàãâéêíóôõú]+){0,4})",
        texto
    )
    if m:
        return m.group(1).strip()

    return ""


def processar_pdf(pdf_path: str) -> Dict:
    texto      = ler_pdf(pdf_path)
    protocolo  = extrair_protocolo(texto)
    unidade    = extrair_unidade(texto)
    data_doc   = extrair_data_documento(texto)
    tipo       = extrair_tipo(texto)
    assunto    = extrair_assunto(texto)
    reclamante = extrair_reclamante(texto, tipo)

    prazo = (data_doc + timedelta(days=PRAZO_DIAS)) if data_doc else None

    return {
        "Protocolo":        protocolo or "NÃO IDENTIFICADO",
        "Tipo":             tipo,
        "Unidade":          unidade,
        "Reclamante":       reclamante,
        "Data Recebimento": data_doc.strftime("%d/%m/%Y") if data_doc else "",
        "Prazo Resposta":   prazo.strftime("%d/%m/%Y") if prazo else "",
        "Assunto":          assunto,
        "Arquivo":          os.path.basename(pdf_path),
        "Status":           "PENDENTE",
        "Observações":      "",
    }

# =========================
# EXCEL: CRIAR OU ATUALIZAR
# =========================
COLUNAS = [
    "Protocolo", "Tipo", "Unidade", "Reclamante", "Data Recebimento",
    "Prazo Resposta", "Assunto", "Arquivo", "Status", "Observações"
]

COR_HEADER   = "2B5597"   # azul prefeitura
COR_OUVIDORIA = "FFE0E0"  # rosa claro
COR_RESPOSTA  = "E0F0E0"  # verde claro
COR_VENCIDO   = "FF4444"  # vermelho
COR_PROXIMO   = "FFA500"  # laranja (vence em 3 dias)

def _borda():
    lado = Side(style="thin", color="CCCCCC")
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def criar_ou_atualizar_excel(novos: List[Dict]) -> None:
    hoje = date.today()

    if os.path.exists(RELATORIO):
        wb = load_workbook(RELATORIO)
        ws = wb.active
        # Carrega protocolos já existentes para não duplicar
        protocolos_existentes = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                protocolos_existentes.add(str(row[0]).strip())
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Ouvidorias"
        protocolos_existentes = set()

        # Cabeçalho
        for col_idx, col_name in enumerate(COLUNAS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
            cell.fill = PatternFill("solid", start_color=COR_HEADER)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _borda()
        ws.row_dimensions[1].height = 22

    # Adiciona apenas novos registros
    linha_inicio = ws.max_row + 1
    adicionados = 0

    for dado in novos:
        proto = str(dado.get("Protocolo", "")).strip()
        if proto and proto != "NÃO IDENTIFICADO" and proto in protocolos_existentes:
            print(f"  Protocolo {proto} já existe no Excel, pulando.")
            continue

        row_num = ws.max_row + 1
        valores = [dado.get(c, "") for c in COLUNAS[:-1]] + [""]  # Observações vazia

        tipo = dado.get("Tipo", "OUVIDORIA")
        cor_linha = COR_OUVIDORIA if tipo == "OUVIDORIA" else COR_RESPOSTA

        # Verifica prazo para colorir
        prazo_str = dado.get("Prazo Resposta", "")
        cor_prazo = None
        if prazo_str:
            try:
                prazo_date = datetime.strptime(prazo_str, "%d/%m/%Y").date()
                diff = (prazo_date - hoje).days
                if diff < 0:
                    cor_prazo = COR_VENCIDO
                elif diff <= 3:
                    cor_prazo = COR_PROXIMO
            except ValueError:
                pass

        for col_idx, valor in enumerate(valores, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=valor)
            cell.font = Font(name="Arial", size=10)
            cell.fill = PatternFill("solid", start_color=cor_linha)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = _borda()

            # Coluna "Prazo Resposta" recebe cor especial se vencido/próximo
            col_nome = COLUNAS[col_idx - 1]
            if col_nome == "Prazo Resposta" and cor_prazo:
                cell.fill = PatternFill("solid", start_color=cor_prazo)
                cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")

        protocolos_existentes.add(proto)
        adicionados += 1

    # Ajusta largura das colunas
    larguras = {
        "Protocolo": 18, "Tipo": 12, "Unidade": 28, "Reclamante": 22,
        "Data Recebimento": 16, "Prazo Resposta": 16,
        "Assunto": 45, "Arquivo": 35, "Status": 12, "Observações": 30
    }
    for col_idx, col_nome in enumerate(COLUNAS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = larguras.get(col_nome, 18)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(RELATORIO)
    print(f"✅ Excel salvo: {RELATORIO} ({adicionados} novo(s) adicionado(s))")

# =========================
# OUTLOOK: IMAP
# =========================
def garantir_pasta_existe(imap, nome: str):
    try:
        folders = [f[2] for f in imap.list_folders()]
        if nome not in folders:
            imap.create_folder(nome)
            print(f"  Pasta IMAP criada: {nome}")
    except Exception as e:
        print(f"  Aviso ao criar pasta '{nome}': {e}")

def obter_message_id(pyz_msg) -> str:
    mid = (
        pyz_msg.get_decoded_header("Message-ID")
        or pyz_msg.get_decoded_header("Message-Id")
        or ""
    )
    mid = str(mid).strip()
    if mid and not (mid.startswith("<") and mid.endswith(">")):
        mid = f"<{mid.strip('<>')}>"
    return mid

def carregar_message_ids_processados(imap) -> set:
    ids = set()
    try:
        imap.select_folder(IMAP_PROCESSADO)
        uids = imap.search(["ALL"])
        if uids:
            fetched = imap.fetch(uids, ["BODY.PEEK[HEADER.FIELDS (MESSAGE-ID)]"])
            for uid, data in fetched.items():
                raw = data.get(b"BODY[HEADER.FIELDS (MESSAGE-ID)]", b"")
                m = re.search(rb"Message-ID:\s*(<[^>]+>)", raw, re.IGNORECASE)
                if m:
                    ids.add(m.group(1).decode().strip())
    except Exception as e:
        print(f"  Aviso ao carregar processados: {e}")
    return ids

# =========================
# PIPELINE PRINCIPAL
# =========================
def processar_emails_api(data_inicial, data_final, log_func=print, somente_nao_lidos=False):
    """Entry point para uso via API — carrega credenciais do config.json."""
    import json as _json
    cfg_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'config.json')
    try:
        with open(cfg_path, encoding='utf-8') as _f:
            _cfg = _json.load(_f)
    except Exception:
        _cfg = {}

    global EMAIL, SENHA_APP, REMETENTE_OUVIDORIA
    EMAIL               = _cfg.get('email',               EMAIL)
    SENHA_APP           = _cfg.get('senha_app',           SENHA_APP)
    REMETENTE_OUVIDORIA = _cfg.get('remetente_ouvidoria', REMETENTE_OUVIDORIA)

    main(data_inicial=data_inicial, data_final=data_final,
         somente_nao_lidos=somente_nao_lidos, log_func=log_func)


def main(data_inicial: Optional[date] = None, data_final: Optional[date] = None,
         somente_nao_lidos: bool = True, log_func=print):
    # Defaults if not provided
    if data_inicial is None:
        data_inicial = date.today()
    if data_final is None:
        data_final = date.today()
    if data_final < data_inicial:
        raise ValueError("Data final não pode ser menor que a data inicial.")

    imap = imapclient.IMAPClient(IMAP_HOST, port=IMAP_PORT, ssl=True)
    imap.login(EMAIL, SENHA_APP)
    imap.select_folder("INBOX")

    for pasta in (IMAP_OUVIDORIAS, IMAP_RESPOSTAS, IMAP_PROCESSADO):
        garantir_pasta_existe(imap, pasta)

    # Carrega Message-IDs já processados
    ids_processados = carregar_message_ids_processados(imap)
    imap.select_folder("INBOX")

    # Critérios de busca
    before_date = data_final + timedelta(days=1)
    criteria = ["SINCE", data_inicial, "BEFORE", before_date]
    if REMETENTE_OUVIDORIA:
        criteria = ["FROM", REMETENTE_OUVIDORIA] + criteria
    if somente_nao_lidos:
        criteria = ["UNSEEN"] + criteria

    uids = imap.search(criteria)
    log_func(f"\n{len(uids)} e-mail(s) encontrado(s) ({data_inicial} → {data_final}).\n")

    todos_dados = []
    pdfs_baixados = 0

    for uid in uids:
        fetched = imap.fetch([uid], ["X-GM-LABELS", "BODY[]"])
        message = pyzmail.PyzMessage.factory(fetched[uid][b"BODY[]"])
        message_id = obter_message_id(message)

        if message_id and message_id in ids_processados:
            log_func(f"[UID {uid}] Já processado, pulando.")
            continue

        # Baixa PDFs do e-mail
        pdf_encontrados = []
        for part in message.mailparts:
            if part.filename and part.filename.lower().endswith(".pdf"):
                nome_arquivo = f"{uid}_{part.filename}"

                # Decide pasta destino provisoriamente (refina após leitura do PDF)
                caminho_tmp = os.path.join(PASTA_OUVIDORIAS, nome_arquivo)
                with open(caminho_tmp, "wb") as f:
                    f.write(part.get_payload())
                pdf_encontrados.append((nome_arquivo, caminho_tmp))
                pdfs_baixados += 1

        if not pdf_encontrados:
            log_func(f"[UID {uid}] Nenhum PDF encontrado, pulando.")
            continue

        # Processa cada PDF
        for nome_arquivo, caminho in pdf_encontrados:
            log_func(f"[UID {uid}] 📄 {nome_arquivo}")
            dados = processar_pdf(caminho)
            dados["EmailUID"] = uid

            log_func(f"  Protocolo  : {dados['Protocolo']}")
            log_func(f"  Tipo       : {dados['Tipo']}")
            log_func(f"  Unidade    : {dados['Unidade']}")
            log_func(f"  Reclamante : {dados.get('Reclamante','')}")
            log_func(f"  Data       : {dados.get('Data Recebimento','')}")
            log_func(f"  Prazo      : {dados['Prazo Resposta']}")

            # Move arquivo para pasta correta
            if dados["Tipo"] == "RESPOSTA":
                destino = os.path.join(PASTA_RESPOSTAS, nome_arquivo)
                imap.add_gmail_labels([uid], [LABEL_RESPOSTAS])
            else:
                destino = os.path.join(PASTA_OUVIDORIAS, nome_arquivo)
                imap.add_gmail_labels([uid], [LABEL_OUVIDORIAS])

            if caminho != destino:
                os.rename(caminho, destino)
                dados["Arquivo"] = os.path.basename(destino)

            todos_dados.append(dados)

        # Marca como processado
        try:
            imap.add_gmail_labels([uid], [LABEL_PROCESSADO])
            ids_processados.add(message_id)
        except Exception as e:
            log_func(f"  Aviso ao marcar processado: {e}")

    imap.logout()

    log_func(f"\nPDFs baixados: {pdfs_baixados}")

    if todos_dados:
        log_func("\nGerando/atualizando Excel...")
        criar_ou_atualizar_excel(todos_dados)

        ouvidorias = [d for d in todos_dados if d["Tipo"] == "OUVIDORIA"]
        respostas  = [d for d in todos_dados if d["Tipo"] == "RESPOSTA"]
        log_func(f"\n{'='*50}")
        log_func(f"📋 Ouvidorias novas : {len(ouvidorias)}")
        log_func(f"📋 Respostas novas  : {len(respostas)}")
        log_func(f"📁 Excel            : {RELATORIO}")
        log_func(f"{'='*50}")
        return {"ouvidorias": len(ouvidorias), "respostas": len(respostas), "pdfs": pdfs_baixados}
    else:
        log_func("\nNenhuma ouvidoria nova para processar.")
        return {"ouvidorias": 0, "respostas": 0, "pdfs": 0}

if __name__ == "__main__":
    # Interactive CLI prompts (kept here so importing this module is side-effect free)
    data_ini_str = input("Baixar ouvidorias a partir de qual data? (YYYY-MM-DD) [ENTER = hoje]: ").strip()
    data_inicial = datetime.strptime(data_ini_str, "%Y-%m-%d").date() if data_ini_str else date.today()

    data_fim_str = input("Até qual data? (YYYY-MM-DD) [ENTER = hoje]: ").strip()
    data_final = datetime.strptime(data_fim_str, "%Y-%m-%d").date() if data_fim_str else date.today()

    if data_final < data_inicial:
        raise SystemExit("Data final não pode ser menor que a data inicial.")

    somente_nao_lidos = input("Apenas NÃO LIDOS? (s/n) [s]: ").strip().lower()
    somente_nao_lidos = somente_nao_lidos in ("", "s", "sim", "y", "yes")

    main(data_inicial=data_inicial, data_final=data_final, somente_nao_lidos=somente_nao_lidos)
