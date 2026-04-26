import os
import re
import difflib
import unicodedata
import smtplib
import threading
import json
import subprocess
from datetime import datetime, date, timedelta
from email.message import EmailMessage
from email.utils import formatdate
from typing import Optional, Dict, List
import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
import subprocess
import platform
from cobrar import executar_cobranca
from tkinter import ttk, filedialog
from tkcalendar import DateEntry   # pip install tkcalendar

import webbrowser
from constantes import CATALOGO_USF, COLUNAS_OUV, normalizar, identificar_unidade

# Dashboard web (Flask — porta 7731)
try:
    from dashboard_server import start_server as _start_dashboard_server
    _DASH_OK = True
except Exception as _dash_e:
    _DASH_OK = False
    _DASH_ERR = str(_dash_e)

# ── Dependências pesadas importadas com tratamento de erro ──────────────────
try:
    import pdfplumber
    import pytesseract
    from pdf2image import convert_from_path
    import imapclient
    import pyzmail
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    # APScheduler (opcional) — import seguro
    try:
        from apscheduler.schedulers.background import BackgroundScheduler
        SCHED_OK = True
        _SCHED_ERR = None
    except Exception as _e:
        BackgroundScheduler = None
        SCHED_OK = False
        _SCHED_ERR = str(_e)
    DEPS_OK = True
except ImportError as _e:
    DEPS_OK = False
    _DEPS_ERRO = str(_e)

# ── Configurações SMTP ──────────────────────────────────────────────────────
SMTP_HOST = "smtp.gmail.com"
SMTP_PORT = 465

# ── Otimizações de Performance ──────────────────────────────────────────────
import concurrent.futures
import time
import threading

# Lock para pytesseract, pois não é thread-safe
tesseract_lock = threading.Lock()

# Compilar regex fora de loops para performance
PATTERNS_PROTOCOLO = [
    # P.O. 003279/2026  — formato Controladoria (4-6 dígitos + ano)
    re.compile(r"P\.?\s*O\.?\s*[:\-]?\s*(\d{3,})\s*[/\-]\s*(\d{4})", re.IGNORECASE),
    # Ref. PO 2278 / Ref: P.O. 2278 — cabeçalho de resposta de unidade
    re.compile(r"Ref\.?\s*[:\-]?\s*P\.?\s*O\.?\s*[\.\-\s]*(\d{3,})", re.IGNORECASE),
    # PO XXXX isolado (sem Ref., sem ano)
    re.compile(r"\bP\.?\s*O\.?\s+(\d{3,})\b", re.IGNORECASE),
    # Campo genérico "Protocolo:"
    re.compile(r"(?:Protocolo|Proc\.?|N[uú]mero)[:\s#]*(\d{4,}[\s/.-]\d{2,4})", re.IGNORECASE),
]

PATTERN_UNIDADE = re.compile(r"PARA\s*:\s*([^\n]{3,80})", re.IGNORECASE)

PATTERNS_DATA = [
    (re.compile(r"(\d{1,2})\s+de\s+([a-záàãâéêíóôõúç]+)\s+(?:de\s+)?(\d{4})", re.IGNORECASE), "%d-%m-%Y"),
    (re.compile(r"(\d{2}[-/]\d{2}[-/]\d{4})", re.IGNORECASE), ["%d-%m-%Y", "%d/%m/%Y"]),
    (re.compile(r"(\d{4}[-/]\d{2}[-/]\d{2})", re.IGNORECASE), ["%Y-%m-%d", "%Y/%m/%d"]),
]

PATTERN_ASSUNTO = re.compile(
    r"OUVIDORIA\s*\n\s*([A-ZÁÀÃÂÉÊÍÓÔÕÚ][^\n]{5,120})\s*\n", re.IGNORECASE)
PATTERN_ASSUNTO_EMENTA = re.compile(
    r"E\s*M\s*E\s*N\s*T\s*A[^\n]*\n\s*([A-ZÁÀÃÂÉÊÍÓÔÕÚ][^\n]{5,120})", re.IGNORECASE)
PATTERN_ASSUNTO_CAT = re.compile(
    r"(ATENDIMENTO\s+INSATISFAT[^\n]{0,60}|FALTA\s*/\s*DEMORA[^\n]{0,80}"
    r"|SERVI[CÇ]O\s+IRREGULAR[^\n]{0,60}|RECLAMAÇÃO[^\n]{0,60})", re.IGNORECASE)
PATTERN_ASSUNTO_ALT = re.compile(r"(?:Trata-se de|Objeto[:\s]+|Assunto[:\s]+)([^\n]{10,150})", re.IGNORECASE)

# Cache para normalização (evita reprocessamento)
_CACHE_NORM = {}

def _norm_cached(texto: str) -> str:
    if texto in _CACHE_NORM:
        return _CACHE_NORM[texto]
    norm = unicodedata.normalize("NFKD", texto)
    norm = "".join(c for c in norm if not unicodedata.combining(c))
    norm = re.sub(r"[^\w\s]", " ", norm)  # Remove pontuação, incluindo parênteses
    norm = re.sub(r"\s+", " ", norm).strip().lower()
    _CACHE_NORM[texto] = norm
    return norm

# =============================================================================
# CONFIG  (salva/carrega em config.json ao lado do script)
# =============================================================================
CONFIG_FILE = os.path.join(os.path.dirname(__file__), "config.json")

CONFIG_PADRAO = {
    "email":               "",
    "senha_app":           "",
    "remetente_ouvidoria": "",
    "pasta_base":          "ouvidorias",
    "prazo_dias":          10,
    "tesseract_path":      r"C:\Program Files\Tesseract-OCR\tesseract.exe",
    "remover_arquivos_depois": True,
    "scheduler_enabled": False,
    "scheduler_interval_min": 60,
}

def carregar_config() -> dict:
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, encoding="utf-8") as f:
                cfg = json.load(f)
            # garante chaves novas
            for k, v in CONFIG_PADRAO.items():
                cfg.setdefault(k, v)
            return cfg
        except Exception:
            pass
    return dict(CONFIG_PADRAO)

def salvar_config(cfg: dict):
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(cfg, f, indent=2, ensure_ascii=False)

# =============================================================================
# FUNÇÕES DE EXTRAÇÃO / PDF
# =============================================================================
def _norm(texto: str) -> str:
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = re.sub(r"[^\w\s]", " ", texto)
    return re.sub(r"\s+", " ", texto).strip().lower()

def identificar_usf(texto: str, threshold: float = 0.6) -> Optional[str]:
    tn = _norm_cached(texto); palavras = tn.split()
    melhor_nome = None; melhor_score = 0.0
    for usf in CATALOGO_USF:
        for alias in usf["aliases"]:
            an = _norm_cached(alias)
            if an in tn:
                return usf["nome"]  # Match exato prioritário
            ap = an.split(); n = len(ap)
            for i in range(max(1, len(palavras) - n + 1)):
                score = difflib.SequenceMatcher(None, an, " ".join(palavras[i:i+n])).ratio()
                if score > melhor_score:
                    melhor_score = score; melhor_nome = usf["nome"]
    return melhor_nome if melhor_score >= threshold else None

def ler_pdf(pdf_path: str, log) -> str:
    texto = ""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for p in pdf.pages:
                t = p.extract_text()
                if t: texto += t + "\n"
    except Exception as e:
        log(f"  Erro pdfplumber: {e}")

    texto = texto.strip()
    sinal_original = verificar_sinais_ouvidoria(texto)

    if len(texto) < 150 or not sinal_original:
        log("  → OCR rápido (1-5 páginas) para complementar texto de PDF")
        try:
            imgs = convert_from_path(pdf_path, first_page=1, last_page=5)  # Limita a 5 páginas para velocidade
            ocr = "".join(pytesseract.image_to_string(img, lang="por") + "\n" for img in imgs).strip()
            if len(ocr) > len(texto):
                texto = ocr
            if not verificar_sinais_ouvidoria(texto):
                log("  → Ainda sem sinais após OCR rápido, tentando OCR completo")
                texto_full = ler_pdf_forcado(pdf_path, log)
                if texto_full and (len(texto_full) > len(texto) or verificar_sinais_ouvidoria(texto_full)):
                    texto = texto_full
        except Exception as e:
            log(f"  Erro OCR: {e}")

    return texto.strip()

def ler_pdf_forcado(pdf_path: str, log) -> str:
    """Força OCR completo em todas as páginas do PDF."""
    log("  → OCR forçado completo...")
    try:
        imgs = convert_from_path(pdf_path)
        ocr = "".join(pytesseract.image_to_string(img, lang="por") + "\n" for img in imgs)
        return ocr.strip()
    except Exception as e:
        log(f"  Erro OCR forçado: {e}")
        return ""

def ler_imagem(img_path: str, log) -> str:
    """Extrai texto de imagem usando OCR com pré-processamento otimizado para velocidade."""
    log(f"  → OCR rápido em imagem: {img_path}")
    try:
        from PIL import Image, ImageEnhance
        log("  → Abrindo imagem...")
        img = Image.open(img_path)
        log(f"  → Imagem: {img.size}")
        
        # Pré-processamento rápido
        if img.mode != 'L':
            img = img.convert('L')  # Grayscale
        enhancer = ImageEnhance.Contrast(img)
        img = enhancer.enhance(1.5)  # Contraste moderado
        
        # Redimensiona para acelerar OCR (menor que antes)
        max_size = (800, 800)
        img.thumbnail(max_size, Image.Resampling.LANCZOS)
        log(f"  → Redimensionada para {img.size}")
        
        # OCR rápido com config otimizada
        log("  → OCR...")
        with tesseract_lock:
            config = '--psm 6 --oem 1 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789áàãâéêíóôõúçÁÀÃÂÉÊÍÓÔÕÚÇ.,;:-/() '
            texto = pytesseract.image_to_string(img, lang="por", config=config)
        log(f"  → OCR concluído: {len(texto)} chars")
        return texto.strip()
    except Exception as e:
        log(f"  → Erro OCR imagem: {e}")
        return ""

def verificar_sinais_ouvidoria(texto: str) -> bool:
    """Verificação rápida se o texto contém sinais de ouvidoria."""
    if not texto:
        return False
    texto_norm = _norm_cached(texto.lower())
    
    # Palavras-chave de ouvidoria (expandido)
    sinais_ouvidoria = [
        "ouvidoria", "protocolo", "manifestação", "reclamação", "denúncia",
        "solicitação", "queixa", "elogio", "sugestão", "p.o.", "p o",
        "proc.", "nº", "numero", "ouv.", "resp.", "resposta", "atendimento",
        "demanda", "requerimento", "petição", "consulta", "informação"
    ]
    
    # Verificar palavras-chave
    for sinal in sinais_ouvidoria:
        if sinal in texto_norm:
            return True
    
    # Verificar nomes de unidades (mais flexível)
    for usf in CATALOGO_USF:
        for alias in usf["aliases"]:
            alias_norm = _norm_cached(alias)
            if alias_norm in texto_norm:
                return True
            # Verificar substrings maiores (ex.: "suzanopolis" em "jardim suzanopolis")
            if len(alias_norm) > 3 and any(alias_norm in palavra for palavra in texto_norm.split() if len(palavra) > 3):
                return True
    
    return False

def extrair_protocolo(texto: str) -> str:
    # Padrão 0 e 1: com grupo (numero, ano) separados
    for pat in PATTERNS_PROTOCOLO[:2]:
        m = pat.search(texto)
        if m:
            grupos = m.groups()
            if len(grupos) == 2:
                return f"P.O. {grupos[0]}/{grupos[1]}"
            return f"P.O. {grupos[0]}"
    # Padrões 2+: grupo único
    for pat in PATTERNS_PROTOCOLO[2:]:
        m = pat.search(texto)
        if m:
            return f"P.O. {m.group(1).replace(' ', '/').strip()}"
    return ""

def extrair_unidade(texto: str) -> str:
    usf = identificar_usf(texto)
    if usf: return usf
    m = PATTERN_UNIDADE.search(texto)
    if m:
        dest = m.group(1).strip()
        if dest and dest.upper() not in ("", "OUVIDORIA", "CONTROLADORIA"):
            return identificar_usf(dest) or dest
    return "NÃO IDENTIFICADA"

def extrair_data_documento(texto: str) -> Optional[date]:
    meses = {"janeiro":1,"fevereiro":2,"março":3,"marco":3,"abril":4,"maio":5,
              "junho":6,"julho":7,"agosto":8,"setembro":9,"outubro":10,"novembro":11,"dezembro":12}
    for pat, fmts in PATTERNS_DATA:
        if isinstance(fmts, str):
            m = pat.search(texto)
            if m:
                dia = int(m.group(1)); mes = meses.get(_norm_cached(m.group(2)), 0); ano = int(m.group(3))
                if mes and 1 <= dia <= 31:
                    try: return date(ano, mes, dia)
                    except ValueError: pass
        else:
            for match in pat.finditer(texto):
                for fmt in fmts:
                    try: return datetime.strptime(match.group(1), fmt).date()
                    except ValueError: pass
    return None

def extrair_tipo(texto: str) -> str:
    inicio = texto[:600]
    # Regra 1: título exclusivo de resposta de unidade
    if re.search(r"Resposta\s+à\s+Manifesta[çc][aã]o\s+de\s+Ouvidoria", inicio, re.IGNORECASE):
        return "RESPOSTA"
    # Regra 2: estrutura da Controladoria → ouvidoria
    if re.search(r"E\s*M\s*E\s*N\s*T\s*A\s+D\s*A\s+D\s*E\s*M\s*A\s*N\s*D\s*A", texto, re.IGNORECASE):
        return "OUVIDORIA"
    if re.search(r"CONTROLADORIA\s+GERAL\s+DO\s+MUNIC", inicio, re.IGNORECASE) and \
       re.search(r"ENCAMINHAMENTO", texto, re.IGNORECASE):
        return "OUVIDORIA"
    # Regra 3: assinatura de gerente/enfermeira → resposta
    if re.search(r"Atenciosamente", texto, re.IGNORECASE) and \
       re.search(r"Enfermeira|Gerente|Coordenadora|Diretor", texto, re.IGNORECASE):
        return "RESPOSTA"
    # Fallback
    tu = texto.upper()
    if any(s in tu for s in ["EM RESPOSTA","VIMOS RESPONDER","RESPONDEMOS",
                               "RETORNO À MANIFESTAÇÃO","RESPOSTA À OUVIDORIA"]):
        return "RESPOSTA"
    return "OUVIDORIA"

def extrair_assunto(texto: str) -> str:
    # Linha destacada entre P.O./OUVIDORIA e EMENTA (estrutura Controladoria)
    m = PATTERN_ASSUNTO.search(texto)
    if m:
        cand = m.group(1).strip()
        if not re.search(r"PREFEITURA|CONTROLADORIA|ESTADO DE", cand, re.IGNORECASE):
            return cand[:120]
    m = PATTERN_ASSUNTO_EMENTA.search(texto)
    if m: return m.group(1).strip()[:120]
    m = PATTERN_ASSUNTO_CAT.search(texto)
    if m: return m.group(1).strip()[:120]
    m = PATTERN_ASSUNTO_ALT.search(texto)
    if m: return m.group(1).strip()[:120]
    return ""

def extrair_reclamante(texto: str, tipo: str = "OUVIDORIA") -> str:
    if tipo == "RESPOSTA":
        m = re.search(
            r"Prezad[ao]\s+Sr[ao]?\.?\s+([A-ZÁÀÃÂÉÊÍÓÔÕÚ][a-záàãâéêíóôõú]+"
            r"(?:\s+[A-ZÁÀÃÂÉÊÍÓÔÕÚ][a-záàãâéêíóôõú]+){1,5})", texto)
        if m: return m.group(1).strip()
    m = re.search(
        r"[Mm]e\s+chamo\s+([A-ZÁÀÃÂÉÊÍÓÔÕÚ][a-záàãâéêíóôõú]+"
        r"(?:\s+[A-ZÁÀÃÂÉÊÍÓÔÕÚ][a-záàãâéêíóôõú]+)*)", texto)
    if m:
        nome = re.split(r'\s+(?:tenho|com|estou|sou)\b', m.group(1), flags=re.IGNORECASE)[0]
        return nome.strip()
    m = re.search(
        r"[Mm]eu\s+nome\s+[eé][:\s]+([A-ZÁÀÃÂÉÊÍÓÔÕÚ][a-záàãâéêíóôõú]+"
        r"(?:\s+[A-ZÁÀÃÂÉÊÍÓÔÕÚ][a-záàãâéêíóôõú]+){0,4})", texto)
    if m: return m.group(1).strip()
    return ""

def processar_documento(file_path: str, data_recebimento: Optional[date], prazo_dias: int, log) -> Dict:
    ext = os.path.splitext(file_path)[1].lower()
    if ext in ('.jpg', '.jpeg', '.png', '.bmp', '.tiff'):
        texto = ler_imagem(file_path, log)
    else:
        texto = ler_pdf(file_path, log)
    
    # Verificação rápida: se não há sinais de ouvidoria, tentar OCR forçado para PDFs
    if not verificar_sinais_ouvidoria(texto):
        if ext not in ('.jpg', '.jpeg', '.png', '.bmp', '.tiff'):
            log("  ⚡ Sem sinais no texto inicial, tentando OCR forçado...")
            texto_ocr = ler_pdf_forcado(file_path, log)
            if texto_ocr and verificar_sinais_ouvidoria(texto_ocr):
                texto = texto_ocr
                log("  ✅ Sinais encontrados no OCR forçado.")
            else:
                log("  ⚡ Sem sinais mesmo no OCR forçado, pulando extração.")
                return {"nao_ouvidoria": True}
        else:
            log("  ⚡ Sem sinais de ouvidoria, pulando extração.")
            return {"nao_ouvidoria": True}
    
    protocolo  = extrair_protocolo(texto)
    unidade    = extrair_unidade(texto)
    data_doc   = extrair_data_documento(texto)
    tipo       = extrair_tipo(texto)
    assunto    = extrair_assunto(texto)
    reclamante = extrair_reclamante(texto, tipo)
    base_prazo = data_recebimento or data_doc
    prazo      = (base_prazo + timedelta(days=prazo_dias)) if base_prazo else None

    observacoes = ""
    if protocolo in ("", "NÃO IDENTIFICADO") or unidade == "NÃO IDENTIFICADA":
        observacoes = "Indeterminada - Revisar extração: sinais de ouvidoria detectados mas dados incompletos. Possível revisão manual necessária."
        log("  ⚠️  Indeterminada: sinais de ouvidoria mas extração incompleta, incluindo com observações.")

    data_rec_str = (data_recebimento or data_doc)
    return {
        "Protocolo":        protocolo or "NÃO IDENTIFICADO",
        "Tipo":             tipo,
        "Unidade":          unidade,
        "Reclamante":       reclamante,
        "Data Recebimento": data_rec_str.strftime("%d/%m/%Y") if data_rec_str else "",
        "Prazo Resposta":   prazo.strftime("%d/%m/%Y") if prazo else "",
        "Assunto":          assunto,
        "Arquivo":          os.path.basename(file_path),
        "Status":           "PENDENTE",
        "Data Respondida":  "",
        "Arquivo Resposta": "",
        "Observações":      observacoes,
    }

# =============================================================================
# ENVIO DE EMAIL COM ANEXO
# =============================================================================
def enviar_email_com_anexo(remetente: str, senha_app: str, destino: str, assunto: str, corpo: str, anexo_path: str, log):
    """Envia um email com anexo."""
    try:
        msg = EmailMessage()
        msg["Subject"] = assunto
        msg["From"] = remetente
        msg["To"] = destino
        msg.set_content(corpo)

        with open(anexo_path, "rb") as f:
            file_data = f.read()
            file_name = os.path.basename(anexo_path)
            msg.add_attachment(file_data, maintype="application", subtype="octet-stream", filename=file_name)

        with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT) as smtp:
            smtp.login(remetente, senha_app)
            smtp.send_message(msg)
        log(f"  ✅ Email enviado para {destino} com anexo {file_name}")
    except Exception as e:
        log(f"  ❌ Erro ao enviar email: {e}")


def remover_arquivos(arquivos: List[str], log):
    """Remove arquivos do disco após o processamento."""
    for caminho in arquivos:
        try:
            if caminho and os.path.exists(caminho):
                os.remove(caminho)
                log(f"  🗑️  Arquivo removido: {caminho}")
        except Exception as e:
            log(f"  ⚠️  Não foi possível remover {caminho}: {e}")

# =============================================================================
# EXCEL
# =============================================================================
COR_HEADER     = "2B5597"
COR_OUV        = "FFF0F0"
COR_RESP       = "F0FFF0"
COR_VENCIDO    = "FF4444"
COR_PROXIMO    = "FFA500"
COR_RESPONDIDA = "D0D0D0"

# Usar colunas de constantes.py e adicionar largura para "Data Última Cobrança"
COLUNAS_RESP = ["Protocolo","Unidade","Data Recebimento","Data Respondida",
                "Assunto","Arquivo Resposta","Observações"]
LARG_OUV  = {"Protocolo":16,"Unidade":28,"Data Recebimento":18,"Prazo Resposta":16,
              "Assunto":48,"Status":14,"Data Respondida":16,"Arquivo":34,"Arquivo Resposta":34,"Observações":28,"Data Última Cobrança":18}
LARG_RESP = {"Protocolo":16,"Unidade":28,"Data Recebimento":18,"Data Respondida":16,
             "Assunto":48,"Arquivo Resposta":34,"Observações":28}

def _borda():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _cab(ws, colunas, larguras):
    for i, col in enumerate(colunas, 1):
        c = ws.cell(row=1, column=i, value=col)
        c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        c.fill      = PatternFill("solid", start_color=COR_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _borda()
    ws.row_dimensions[1].height = 24
    for i, col in enumerate(colunas, 1):
        ws.column_dimensions[get_column_letter(i)].width = larguras.get(col, 18)
    ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions

def _linha(ws, row_num, colunas, dado, cor_base, hoje, col_prazo="Prazo Resposta"):
    for i, col in enumerate(colunas, 1):
        valor = dado.get(col, "")
        c = ws.cell(row=row_num, column=i, value=valor)
        c.font      = Font(name="Arial", size=10)
        c.fill      = PatternFill("solid", start_color=cor_base)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border    = _borda()
        if col == col_prazo and valor:
            try:
                diff = (datetime.strptime(str(valor), "%d/%m/%Y").date() - hoje).days
                cp   = COR_VENCIDO if diff < 0 else (COR_PROXIMO if diff <= 3 else None)
                if cp:
                    c.fill = PatternFill("solid", start_color=cp)
                    c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            except ValueError: pass
    ws.row_dimensions[row_num].height = 34

def _mapa(ws):
    return {str(row[0].value).strip(): row[0].row
            for row in ws.iter_rows(min_row=2) if row[0].value}

def atualizar_excel(relatorio, novas_ouvidorias, novas_respostas, log):
    hoje = date.today()
    wb   = load_workbook(relatorio) if os.path.exists(relatorio) else Workbook()
    if not os.path.exists(relatorio):
        wb.remove(wb.active)

    for nome, colunas, larguras in [("Ouvidorias", COLUNAS_OUV, LARG_OUV),
                                     ("Respostas",  COLUNAS_RESP, LARG_RESP)]:
        if nome not in wb.sheetnames:
            ws = wb.create_sheet(nome); _cab(ws, colunas, larguras)

    ws_ouv = wb["Ouvidorias"]; ws_res = wb["Respostas"]
    mapa_ouv = _mapa(ws_ouv);  mapa_res = _mapa(ws_res)

    col_status    = COLUNAS_OUV.index("Status") + 1
    col_data_resp = COLUNAS_OUV.index("Data Respondida") + 1
    col_arq_resp  = COLUNAS_OUV.index("Arquivo Resposta") + 1

    add_ouv = add_res = 0
    for d in novas_ouvidorias:
        proto = str(d.get("Protocolo","")).strip()
        if proto and proto != "NÃO IDENTIFICADO" and proto in mapa_ouv:
            log(f"  Ouvidoria {proto} já existe, pulando."); continue
        _linha(ws_ouv, ws_ouv.max_row+1, COLUNAS_OUV, d, COR_OUV, hoje)
        mapa_ouv[proto] = ws_ouv.max_row; add_ouv += 1

    for d in novas_respostas:
        proto     = str(d.get("Protocolo","")).strip()
        data_resp = d.get("Data Recebimento","") or d.get("Data Documento","")
        arq_resp  = d.get("Arquivo","")
        if proto in mapa_ouv:
            ln = mapa_ouv[proto]
            c  = ws_ouv.cell(row=ln, column=col_status)
            c.value = "RESPONDIDA"; c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            c.fill  = PatternFill("solid", start_color="4CAF50")
            ws_ouv.cell(row=ln, column=col_data_resp).value = data_resp
            ws_ouv.cell(row=ln, column=col_arq_resp).value  = arq_resp
            for ci in range(1, len(COLUNAS_OUV)+1):
                if ci != col_status:
                    ws_ouv.cell(row=ln, column=ci).fill = PatternFill("solid", start_color=COR_RESPONDIDA)
            log(f"  ✅ Ouvidoria {proto} → RESPONDIDA")
        else:
            log(f"  ⚠️  Resposta {proto}: ouvidoria original não encontrada")
        if proto in mapa_res:
            log(f"  Resposta {proto} já existe, pulando."); continue
        dr = {"Protocolo":proto,"Unidade":d.get("Unidade",""),
              "Data Recebimento":d.get("Data Recebimento",""),
              "Data Respondida":data_resp,"Assunto":d.get("Assunto",""),
              "Arquivo Resposta":arq_resp,"Observações":""}
        _linha(ws_res, ws_res.max_row+1, COLUNAS_RESP, dr, COR_RESP, hoje, col_prazo="")
        mapa_res[proto] = ws_res.max_row; add_res += 1

    for ws, colunas, larguras in [(ws_ouv,COLUNAS_OUV,LARG_OUV),(ws_res,COLUNAS_RESP,LARG_RESP)]:
        for i, col in enumerate(colunas, 1):
            ws.column_dimensions[get_column_letter(i)].width = larguras.get(col,18)
        ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions

    wb.save(relatorio)
    log(f"✅ Excel salvo: {relatorio} | +{add_ouv} ouvidoria(s) | +{add_res} resposta(s)")
    return add_ouv, add_res

# =============================================================================
# PIPELINE GMAIL
# =============================================================================
def garantir_label(imap, nome):
    try:
        if nome not in [f[2] for f in imap.list_folders()]:
            imap.create_folder(nome); 
    except Exception: pass

def get_mid(msg):
    mid = msg.get_decoded_header("Message-ID") or msg.get_decoded_header("Message-Id") or ""
    mid = str(mid).strip()
    if mid and not (mid.startswith("<") and mid.endswith(">")):
        mid = f"<{mid.strip('<>')}>"
    return mid

def carregar_processados(imap, label):
    ids = set()
    try:
        imap.select_folder(label)
        uids = imap.search(["ALL"])
        if uids:
            fetched = imap.fetch(uids, ["BODY.PEEK[HEADER.FIELDS (MESSAGE-ID)]"])
            for uid, data in fetched.items():
                raw = data.get(b"BODY[HEADER.FIELDS (MESSAGE-ID)]", b"")
                m = re.search(rb"Message-ID:\s*(<[^>]+>)", raw, re.IGNORECASE)
                if m: ids.add(m.group(1).decode().strip())
    except Exception: pass
    return ids

def executar_pipeline(cfg: dict, data_ini: date, data_fim: date,
                       apenas_nao_lidos: bool, log):
    """Roda o pipeline completo. Chamado em thread separada pela GUI."""
    # ── Métricas de Monitoramento ────────────────────────────────────────────
    start_time = time.time()
    metrics = {
        "emails_encontrados": 0,
        "emails_processados": 0,
        "documentos_baixados": 0,
        "ouvidorias_novas": 0,
        "respostas_novas": 0,
        "nao_ouvidorias": 0,
        "erros_pdf": 0,
        "tempo_conexao": 0,
        "tempo_processamento": 0,
        "tempo_excel": 0,
    }
    EMAIL           = cfg["email"]
    SENHA_APP       = cfg["senha_app"]
    REMETENTE       = cfg["remetente_ouvidoria"]
    PASTA_BASE      = cfg["pasta_base"]
    PRAZO_DIAS      = int(cfg["prazo_dias"])
    TESSERACT_PATH  = cfg.get("tesseract_path","")
    RELATORIO       = os.path.join(PASTA_BASE, "ouvidorias.xlsx")
    PASTA_OUV       = os.path.join(PASTA_BASE, "ouvidorias")
    PASTA_RES       = os.path.join(PASTA_BASE, "respostas")
    LABEL_OUV       = "Ouvidorias/Ouvidorias"
    LABEL_RES       = "Ouvidorias/Respostas"
    LABEL_PROC      = "Ouvidorias/Processado"

    os.makedirs(PASTA_OUV, exist_ok=True)
    os.makedirs(PASTA_RES, exist_ok=True)

    if TESSERACT_PATH and os.path.exists(TESSERACT_PATH):
        pytesseract.pytesseract.tesseract_cmd = TESSERACT_PATH

    log(f"Conectando ao Gmail ({EMAIL})...")
    conn_start = time.time()
    try:
        imap = imapclient.IMAPClient("imap.gmail.com", port=993, ssl=True)
        imap.login(EMAIL, SENHA_APP)
    except Exception as e:
        log(f"❌ Erro de login: {e}")
        log("Verifique o e-mail e a senha de app nas Configurações.")
        return
    metrics["tempo_conexao"] = time.time() - conn_start

    imap.select_folder("INBOX")
    log("  Conectado ✅")

    for lbl in (LABEL_OUV, LABEL_RES, LABEL_PROC, "Ouvidorias/Não Ouvidoria"):
        garantir_label(imap, lbl)

    ids_proc = carregar_processados(imap, LABEL_PROC)
    imap.select_folder("INBOX")

    criteria = ["SINCE", data_ini, "BEFORE", data_fim + timedelta(days=1)]
    if REMETENTE: criteria = ["FROM", REMETENTE] + criteria
    if apenas_nao_lidos: criteria = ["UNSEEN"] + criteria

    def eh_email_auto_enviado(msg, meu_email):
        if not msg: return False
        try:
            remetentes = [addr.lower() for _, addr in msg.get_addresses('from')]
        except Exception:
            remetentes = []
        try:
            sender = [addr.lower() for _, addr in msg.get_addresses('sender')]
        except Exception:
            sender = []
        return meu_email.lower() in remetentes or meu_email.lower() in sender

    uids = imap.search(criteria)
    metrics["emails_encontrados"] = len(uids)
    log(f"\n{len(uids)} e-mail(s) encontrado(s) ({data_ini} → {data_fim})\n")

    novas_ouv: List[Dict] = []
    novas_res: List[Dict] = []
    pdfs = 0
    proc_start = time.time()

    for uid in uids:
        arquivos_para_remover = []
        fetched    = imap.fetch([uid], ["X-GM-LABELS","INTERNALDATE","BODY[]"])
        message    = pyzmail.PyzMessage.factory(fetched[uid][b"BODY[]"])
        message_id = get_mid(message)

        if message_id and message_id in ids_proc:
            log(f"[UID {uid}] Já processado, pulando.")
            continue

        if eh_email_auto_enviado(message, EMAIL):
            log(f"[UID {uid}] E-mail autoenviado ({EMAIL}), pulando para evitar redundância.")
            try:
                imap.add_gmail_labels([uid], [LABEL_PROC])
                if message_id:
                    ids_proc.add(message_id)
            except Exception as e:
                log(f"  Aviso ao marcar autoenviado: {e}")
            continue

        dt_raw = fetched[uid].get(b"INTERNALDATE")
        data_recebimento: Optional[date] = None
        if dt_raw and hasattr(dt_raw, "date"):
            data_recebimento = dt_raw.date()

        doc_list = []
        for part in message.mailparts:
            if part.filename:
                nome_arquivo = part.filename.lower()
                if "outlook" in nome_arquivo:
                    log(f"[UID {uid}] Ignorando anexo de assinatura virtual: {part.filename}")
                    continue
                if nome_arquivo.endswith((".pdf", ".jpg", ".jpeg", ".png", ".bmp", ".tiff")):
                    nome    = f"{uid}_{part.filename}"
                    caminho = os.path.join(PASTA_OUV, nome)
                    with open(caminho, "wb") as f:
                        f.write(part.get_payload())
                    doc_list.append((nome, caminho)); pdfs += 1

        metrics["documentos_baixados"] += len(doc_list)

        if not doc_list:
            log(f"[UID {uid}] Sem documento (PDF/imagem), pulando."); continue

        # Processar documentos em paralelo para melhor performance
        dados_list = []
        def processar_documento_paralelo(doc_info):
            nome, caminho = doc_info
            log(f"[UID {uid}] 📄 {nome}")
            dados = processar_documento(caminho, data_recebimento, PRAZO_DIAS, log)
            dados["nome"] = nome
            dados["caminho"] = caminho
            dados_list.append(dados)
            return dados

        with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
            futures = [executor.submit(processar_documento_paralelo, doc) for doc in doc_list]
            for future in concurrent.futures.as_completed(futures):
                try:
                    result = future.result()
                    # Log se necessário
                except Exception as e:
                    log(f"  Erro no processamento paralelo: {e}")

        # Verificar se há múltiplos anexos e pelo menos uma ouvidoria
        ouvidorias_no_email = [d for d in dados_list if not d.get("nao_ouvidoria")]
        if len(doc_list) > 1 and ouvidorias_no_email:
            log(f"[UID {uid}] 📧 Múltiplos anexos com ouvidoria detectada, reenviando anexos ouvidoria separadamente.")
            for dados in ouvidorias_no_email:
                caminho = dados["caminho"]
                nome = dados["nome"]
                assunto = f"Ouvidoria Separada - {nome}"
                corpo = f"Anexo ouvidoria separado do email UID {uid}."
                enviar_email_com_anexo(EMAIL, SENHA_APP, EMAIL, assunto, corpo, caminho, log)
                arquivos_para_remover.append(caminho)
            # Marcar como processado sem adicionar ao Excel
            try:
                imap.add_gmail_labels([uid], [LABEL_PROC])
                ids_proc.add(message_id)
            except Exception as e:
                log(f"  Aviso ao marcar processado: {e}")
            if cfg.get("remover_arquivos_depois", False):
                remover_arquivos(arquivos_para_remover, log)
            continue  # Pula o processamento normal

        # Processamento normal
        for dados in dados_list:
            # Se não é ouvidoria pela verificação rápida, marcar diretamente
            if dados.get("nao_ouvidoria"):
                log(f"  ⚠️  Não identificado como ouvidoria (verificação rápida), marcando como 'Não Ouvidoria'.")
                metrics["nao_ouvidorias"] += 1
                try:
                    imap.add_gmail_labels([uid], ["Ouvidorias/Não Ouvidoria"])
                    imap.add_gmail_labels([uid], [LABEL_PROC])
                    ids_proc.add(message_id)
                except Exception as e:
                    log(f"  Aviso ao marcar não ouvidoria: {e}")
                continue
            
            dados["EmailUID"] = uid
            log(f"  Protocolo : {dados['Protocolo']}")
            log(f"  Tipo      : {dados['Tipo']}")
            log(f"  Unidade   : {dados['Unidade']}")
            log(f"  Recebido  : {dados['Data Recebimento']}  Prazo: {dados['Prazo Resposta']}")

            if dados["Tipo"] == "RESPOSTA":
                destino = os.path.join(PASTA_RES, dados["nome"])
                novas_res.append(dados)
            else:
                destino = os.path.join(PASTA_OUV, dados["nome"])
                novas_ouv.append(dados)

            caminho = dados["caminho"]
            if caminho != destino:
                os.rename(caminho, destino)
                dados["Arquivo"] = os.path.basename(destino)
            arquivos_para_remover.append(destino)

        if cfg.get("remover_arquivos_depois", False):
            remover_arquivos(arquivos_para_remover, log)

        try:
            imap.add_gmail_labels([uid], [LABEL_PROC])
            ids_proc.add(message_id)
        except Exception as e:
            log(f"  Aviso ao marcar processado: {e}")

    metrics["tempo_processamento"] = time.time() - proc_start
    metrics["emails_processados"] = len([uid for uid in uids if not (message_id and message_id in ids_proc)])  # Aproximado

    imap.logout()
    log(f"\nDocumentos baixados: {pdfs}")

    if novas_ouv or novas_res:
        excel_start = time.time()
        log("\nAtualizando Excel...")
        add_ouv, add_res = atualizar_excel(RELATORIO, novas_ouv, novas_res, log)
        metrics["tempo_excel"] = time.time() - excel_start
        metrics["ouvidorias_novas"] = add_ouv
        metrics["respostas_novas"] = add_res
        log(f"\n{'='*48}")
        log(f"📋 Ouvidorias novas : {add_ouv}")
        log(f"📋 Respostas novas  : {add_res}")
        log(f"📁 Excel            : {RELATORIO}")
        log(f"{'='*48}")
    else:
        log("\nNenhuma ouvidoria nova para processar.")

    # ── Log de Métricas ─────────────────────────────────────────────────────
    total_time = time.time() - start_time
    log(f"\n📊 Métricas de Performance:")
    log(f"  Tempo total: {total_time:.2f}s")
    log(f"  Tempo conexão IMAP: {metrics['tempo_conexao']:.2f}s")
    log(f"  Tempo processamento: {metrics['tempo_processamento']:.2f}s")
    log(f"  Tempo Excel: {metrics['tempo_excel']:.2f}s")
    log(f"  E-mails encontrados: {metrics['emails_encontrados']}")
    log(f"  E-mails processados: {metrics['emails_processados']}")
    log(f"  PDFs baixados: {metrics['documentos_baixados']}")
    log(f"  Ouvidorias novas: {metrics['ouvidorias_novas']}")
    log(f"  Respostas novas: {metrics['respostas_novas']}")
    log(f"  Não Ouvidorias: {metrics['nao_ouvidorias']}")
    log(f"  Taxa de processamento: {metrics['emails_processados']/total_time:.2f} e-mails/s" if total_time > 0 else "  Taxa: N/A")

# =============================================================================
# INTERFACE TKINTER
# =============================================================================
COR_AZUL    = "#2B5597"
COR_BRANCO  = "#FFFFFF"
COR_CINZA   = "#F4F6FA"
COR_TEXTO   = "#1A1A2E"
COR_VERDE   = "#4CAF50"
COR_LARANJA = "#FF6B35"
COR_BORDA   = "#D0D7E8"

def _fonte(size=10, bold=False):
    return ("Segoe UI", size, "bold" if bold else "normal")

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.cfg = carregar_config()
        self.rodando = False

        self.title("Bot de Ouvidorias — Prefeitura de Suzano")
        self.geometry("860x680")
        self.minsize(780, 580)
        self.configure(bg=COR_CINZA)
        self.resizable(True, True)

        self._build_ui()
        self._verificar_deps()

        # Scheduler (inicializado quando solicitado)
        self.scheduler = None
        self._scheduler_job = None
        self._setup_scheduler_state()
        # Auto-iniciar se estava habilitado
        if self.cfg.get("scheduler_enabled", False):
            if self._start_scheduler() and hasattr(self, 'btn_sched_toggle'):
                self.btn_sched_toggle.configure(text="Parar Agendador")

        # Inicia servidor do dashboard em background
        if _DASH_OK:
            self._dash_port = _start_dashboard_server(port=7731)
        else:
            self._dash_port = None

    # ── layout ──────────────────────────────────────────────────────────────
    def _build_ui(self):
        # ── cabeçalho ──
        header = tk.Frame(self, bg=COR_AZUL, height=64)
        header.pack(fill="x")
        tk.Label(header, text="🏛  Bot de Ouvidorias", font=_fonte(16, True),
                 bg=COR_AZUL, fg=COR_BRANCO).pack(side="left", padx=20, pady=14)
        tk.Label(header, text="Prefeitura Municipal de Suzano",
                 font=_fonte(9), bg=COR_AZUL, fg="#B0C4DE").pack(side="left", pady=18)

        # ── notebook (abas) ──
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("TNotebook",        background=COR_CINZA, borderwidth=0)
        style.configure("TNotebook.Tab",    font=_fonte(10), padding=[16,6],
                        background=COR_BORDA, foreground=COR_TEXTO)
        style.map("TNotebook.Tab",          background=[("selected", COR_AZUL)],
                                            foreground=[("selected", COR_BRANCO)])
        style.configure("TEntry",           font=_fonte(10))
        style.configure("TCombobox",        font=_fonte(10))

        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True, padx=12, pady=8)

        self._aba_processar(nb)
        self._aba_planilha(nb)
        self._aba_configuracoes(nb)

        # ── barra de status ──
        bar = tk.Frame(self, bg=COR_BORDA, height=28)
        bar.pack(fill="x", side="bottom")
        self.lbl_status = tk.Label(bar, text="Pronto.", font=_fonte(9),
                                   bg=COR_BORDA, fg=COR_TEXTO)
        self.lbl_status.pack(side="left", padx=10)

    # ── aba Processar ────────────────────────────────────────────────────────
    def _aba_processar(self, nb):
        frm = tk.Frame(nb, bg=COR_CINZA)
        nb.add(frm, text="  ▶  Processar  ")

        # painel de filtros
        pnl = tk.LabelFrame(frm, text=" Filtros de busca ", font=_fonte(10, True),
                             bg=COR_CINZA, fg=COR_AZUL, bd=1, relief="groove")
        pnl.pack(fill="x", padx=16, pady=(14,6))

        # datas
        tk.Label(pnl, text="Data inicial:", font=_fonte(10), bg=COR_CINZA).grid(
            row=0, column=0, padx=12, pady=10, sticky="e")
        self.cal_ini = DateEntry(pnl, width=13, font=_fonte(10),
                                 date_pattern="yyyy-mm-dd",
                                 background=COR_AZUL, foreground=COR_BRANCO)
        self.cal_ini.grid(row=0, column=1, padx=6, pady=10, sticky="w")
        self.cal_ini.set_date(date.today())

        tk.Label(pnl, text="Data final:", font=_fonte(10), bg=COR_CINZA).grid(
            row=0, column=2, padx=12, pady=10, sticky="e")
        self.cal_fim = DateEntry(pnl, width=13, font=_fonte(10),
                                 date_pattern="yyyy-mm-dd",
                                 background=COR_AZUL, foreground=COR_BRANCO)
        self.cal_fim.grid(row=0, column=3, padx=6, pady=10, sticky="w")
        self.cal_fim.set_date(date.today())

        # apenas não lidos
        self.var_nao_lidos = tk.BooleanVar(value=False)
        tk.Checkbutton(pnl, text="Apenas e-mails não lidos",
                       variable=self.var_nao_lidos, font=_fonte(10),
                       bg=COR_CINZA, fg=COR_TEXTO,
                       activebackground=COR_CINZA).grid(
            row=0, column=4, padx=20, pady=10, sticky="w")

        # botões
        btn_frm = tk.Frame(frm, bg=COR_CINZA)
        btn_frm.pack(fill="x", padx=16, pady=4)

        self.btn_run = tk.Button(btn_frm, text="▶  Processar Ouvidorias",
                                 font=_fonte(11, True), bg=COR_VERDE, fg=COR_BRANCO,
                                 activebackground="#388E3C", cursor="hand2",
                                 relief="flat", padx=18, pady=8,
                                 command=self._iniciar)
        self.btn_run.pack(side="left", padx=(0,10))

        tk.Button(btn_frm, text="📂  Abrir Excel",
                  font=_fonte(10), bg=COR_AZUL, fg=COR_BRANCO,
                  activebackground="#1A3A7A", cursor="hand2",
                  relief="flat", padx=14, pady=8,
                  command=self._abrir_excel).pack(side="left", padx=(0,10))

        tk.Button(btn_frm, text="🌐  Dashboard Web",
                  font=_fonte(10), bg="#1A5C3A", fg=COR_BRANCO,
                  activebackground="#0F3D26", cursor="hand2",
                  relief="flat", padx=14, pady=8,
                  command=self._abrir_dashboard).pack(side="left", padx=(0,10))

        tk.Button(btn_frm, text="🗑  Limpar Log",
                  font=_fonte(10), bg="#888", fg=COR_BRANCO,
                  activebackground="#555", cursor="hand2",
                  relief="flat", padx=14, pady=8,
                  command=self._limpar_log).pack(side="left")

        # barra de progresso
        self.progress = ttk.Progressbar(frm, mode="indeterminate", length=300)
        self.progress.pack(fill="x", padx=16, pady=(4,0))

        # log
        log_frm = tk.LabelFrame(frm, text=" Log de execução ", font=_fonte(10, True),
                                 bg=COR_CINZA, fg=COR_AZUL, bd=1, relief="groove")
        log_frm.pack(fill="both", expand=True, padx=16, pady=10)

        self.txt_log = tk.Text(log_frm, font=("Consolas", 9), bg="#0D1117",
                               fg="#C9D1D9", insertbackground=COR_BRANCO,
                               relief="flat", wrap="word", state="disabled",
                               padx=10, pady=8)
        sc = ttk.Scrollbar(log_frm, command=self.txt_log.yview)
        self.txt_log.configure(yscrollcommand=sc.set)
        sc.pack(side="right", fill="y")
        self.txt_log.pack(fill="both", expand=True)

        # tags de cor no log
        self.txt_log.tag_configure("ok",    foreground="#58D68D")
        self.txt_log.tag_configure("erro",  foreground="#EC7063")
        self.txt_log.tag_configure("aviso", foreground="#F7DC6F")
        self.txt_log.tag_configure("info",  foreground="#85C1E9")

    # ── aba Configurações ────────────────────────────────────────────────────
    def _aba_configuracoes(self, nb):
        frm = tk.Frame(nb, bg=COR_CINZA)
        nb.add(frm, text="  ⚙  Configurações  ")

        pnl = tk.LabelFrame(frm, text=" Conta Gmail ", font=_fonte(10, True),
                             bg=COR_CINZA, fg=COR_AZUL, bd=1, relief="groove")
        pnl.pack(fill="x", padx=20, pady=(18,8))

        campos = [
            ("E-mail Gmail:",          "email",               False),
            ("Senha de App (16 dig.):", "senha_app",           True),
            ("Remetente (filtro):",     "remetente_ouvidoria", False),
        ]
        self._entries = {}
        for i, (label, key, oculto) in enumerate(campos):
            tk.Label(pnl, text=label, font=_fonte(10), bg=COR_CINZA, anchor="e").grid(
                row=i, column=0, padx=14, pady=8, sticky="e")
            var = tk.StringVar(value=self.cfg.get(key,""))
            e = ttk.Entry(pnl, textvariable=var, width=42,
                          show="●" if oculto else "")
            e.grid(row=i, column=1, padx=8, pady=8, sticky="w")
            self._entries[key] = var

        pnl2 = tk.LabelFrame(frm, text=" Pastas e prazo ", font=_fonte(10, True),
                              bg=COR_CINZA, fg=COR_AZUL, bd=1, relief="groove")
        pnl2.pack(fill="x", padx=20, pady=8)

        campos2 = [
            ("Pasta de saída (local):", "pasta_base",    False),
            ("Prazo (dias):",           "prazo_dias",    False),
            ("Caminho Tesseract.exe:",  "tesseract_path",False),
        ]
        for i, (label, key, _) in enumerate(campos2):
            tk.Label(pnl2, text=label, font=_fonte(10), bg=COR_CINZA, anchor="e").grid(
                row=i, column=0, padx=14, pady=8, sticky="e")
            var = tk.StringVar(value=str(self.cfg.get(key,"")))
            e = ttk.Entry(pnl2, textvariable=var, width=42)
            e.grid(row=i, column=1, padx=8, pady=8, sticky="w")
            self._entries[key] = var
            if key in ("pasta_base","tesseract_path"):
                tk.Button(pnl2, text="📁", font=_fonte(10), relief="flat",
                          cursor="hand2", bg=COR_CINZA,
                          command=lambda k=key: self._browse(k)).grid(
                    row=i, column=2, padx=4)

                # Agendador
                sched_pnl = tk.LabelFrame(frm, text=" Agendador ", font=_fonte(10, True),
                                                                 bg=COR_CINZA, fg=COR_AZUL, bd=1, relief="groove")
                sched_pnl.pack(fill="x", padx=20, pady=(8,8))

                self.var_sched_enabled = tk.BooleanVar(value=bool(self.cfg.get("scheduler_enabled", False)))
                tk.Checkbutton(sched_pnl, text="Ativar agendamento automático",
                                             variable=self.var_sched_enabled, font=_fonte(10), bg=COR_CINZA,
                                             activebackground=COR_CINZA).grid(row=0, column=0, padx=12, pady=8, sticky="w")

                tk.Label(sched_pnl, text="Intervalo (min):", font=_fonte(10), bg=COR_CINZA).grid(row=0, column=1, padx=8, pady=8, sticky="e")
                self.var_sched_interval = tk.StringVar(value=str(self.cfg.get("scheduler_interval_min", 60)))
                ttk.Entry(sched_pnl, textvariable=self.var_sched_interval, width=8).grid(row=0, column=2, padx=6, pady=8, sticky="w")

                self.btn_sched_toggle = tk.Button(sched_pnl, text=("Parar Agendador" if self.var_sched_enabled.get() else "Iniciar Agendador"),
                                                                                    font=_fonte(10), bg=COR_AZUL, fg=COR_BRANCO, relief="flat",
                                                                                    command=lambda: self._toggle_scheduler(self.var_sched_enabled, self.var_sched_interval, self.btn_sched_toggle))
                self.btn_sched_toggle.grid(row=0, column=3, padx=12, pady=8)

                tk.Label(sched_pnl, text="(O job rodará em background e registrará no log)", font=_fonte(9), bg=COR_CINZA, fg="#666").grid(row=1, column=0, columnspan=4, sticky="w", padx=12)

                tk.Button(frm, text="💾  Salvar Configurações",
                                    font=_fonte(11, True), bg=COR_AZUL, fg=COR_BRANCO,
                                    activebackground="#1A3A7A", cursor="hand2",
                                    relief="flat", padx=18, pady=8,
                                    command=self._salvar_cfg).pack(pady=16)

        self.lbl_cfg_ok = tk.Label(frm, text="", font=_fonte(10), bg=COR_CINZA)
        self.lbl_cfg_ok.pack()

    # ── ações ────────────────────────────────────────────────────────────────
    def _log(self, msg: str):
        """Escreve no log de forma thread-safe."""
        def _write():
            self.txt_log.configure(state="normal")
            # detecta tag
            tag = "info"
            if any(x in msg for x in ["✅","OK","sucesso","Conectado"]):  tag = "ok"
            elif any(x in msg for x in ["❌","Erro","erro","ERRO"]):       tag = "erro"
            elif any(x in msg for x in ["⚠️","Aviso","aviso","pulando"]): tag = "aviso"
            self.txt_log.insert("end", msg+"\n", tag)
            self.txt_log.see("end")
            self.txt_log.configure(state="disabled")
            self.lbl_status.configure(text=msg[:80])
        self.after(0, _write)

    def _limpar_log(self):
        self.txt_log.configure(state="normal")
        self.txt_log.delete("1.0","end")
        self.txt_log.configure(state="disabled")

    def _browse(self, key):
        if key == "pasta_base":
            path = filedialog.askdirectory(title="Selecione a pasta de saída")
        else:
            path = filedialog.askopenfilename(title="Selecione o tesseract.exe",
                                              filetypes=[("exe","*.exe"),("Todos","*.*")])
        if path:
            self._entries[key].set(path)

    def _salvar_cfg(self):
        for key, var in self._entries.items():
            self.cfg[key] = var.get().strip()
        salvar_config(self.cfg)
        self.lbl_cfg_ok.configure(text="✅ Configurações salvas!", fg=COR_VERDE)
        self.after(3000, lambda: self.lbl_cfg_ok.configure(text=""))

    def _abrir_excel(self):
        pasta  = self.cfg.get("pasta_base","ouvidorias")
        relat  = os.path.join(pasta, "ouvidorias.xlsx")
        if os.path.exists(relat):
            os.startfile(relat)
        else:
            messagebox.showinfo("Arquivo não encontrado",
                                f"Ainda não há planilha em:\n{relat}\n\n"
                                "Execute o processamento primeiro.")

    def _abrir_dashboard(self):
        if not _DASH_OK:
            messagebox.showerror("Dashboard indisponível",
                                 f"Não foi possível iniciar o servidor web.\n\n{_DASH_ERR}\n\n"
                                 "Instale com: pip install flask flask-cors")
            return
        port = getattr(self, '_dash_port', 7731) or 7731
        url  = f"http://localhost:{port}"
        webbrowser.open(url)

    def _iniciar(self):
        if not DEPS_OK:
            messagebox.showerror("Dependências ausentes",
                                 f"Módulo não instalado: {_DEPS_ERRO}\n\n"
                                 "Execute no terminal:\n"
                                 "pip install pdfplumber pytesseract pdf2image "
                                 "imapclient pyzmail36 openpyxl pandas tkcalendar")
            return
        if self.rodando:
            messagebox.showwarning("Aguarde", "Processamento em andamento...")
            return

        # Coleta parâmetros
        try:
            data_ini = self.cal_ini.get_date()
            data_fim = self.cal_fim.get_date()
        except Exception:
            messagebox.showerror("Data inválida", "Verifique as datas selecionadas.")
            return

        if data_fim < data_ini:
            messagebox.showerror("Data inválida", "A data final deve ser ≥ data inicial.")
            return

        # Salva config atual antes de rodar
        for key, var in self._entries.items():
            self.cfg[key] = var.get().strip()
        salvar_config(self.cfg)

        self.rodando = True
        self.btn_run.configure(state="disabled", text="⏳  Processando...")
        self.progress.start(12)
        self._log(f"\n{'─'*50}")
        self._log(f"Iniciando — {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")

        apenas_nao_lidos = self.var_nao_lidos.get()

        def _run():
            try:
                executar_pipeline(self.cfg, data_ini, data_fim, apenas_nao_lidos, self._log)
            except Exception as e:
                self._log(f"❌ Erro inesperado: {e}")
            finally:
                self.after(0, self._finalizar)

        threading.Thread(target=_run, daemon=True).start()

    def _finalizar(self):
        self.rodando = False
        self.progress.stop()
        self.btn_run.configure(state="normal", text="▶  Processar Ouvidorias")
        self.lbl_status.configure(text="Processamento concluído.")

    def _verificar_deps(self):
        if not DEPS_OK:
            self._log(f"⚠️  Dependência ausente: {_DEPS_ERRO}")
            self._log("Execute: pip install pdfplumber pytesseract pdf2image "
                      "imapclient pyzmail36 openpyxl pandas tkcalendar")
        else:
            self._log("✅ Todas as dependências carregadas.")
            # Verificar Tesseract
            try:
                import pytesseract
                version = pytesseract.get_tesseract_version()
                self._log(f"✅ Tesseract detectado: {version}")
            except Exception as e:
                self._log(f"⚠️  Tesseract não encontrado: {e}")
                self._log("Instale Tesseract-OCR e certifique-se de que está no PATH ou configure o caminho em Configurações.")
            self._log("Configure sua senha de app na aba ⚙ Configurações e clique em Processar.")

    # ── Scheduler helpers ─────────────────────────────────────────────────
    def _setup_scheduler_state(self):
        # carrega estado salvo
        enabled = bool(self.cfg.get("scheduler_enabled", False))
        interval = int(self.cfg.get("scheduler_interval_min", 60))
        self.cfg.setdefault("scheduler_enabled", enabled)
        self.cfg.setdefault("scheduler_interval_min", interval)

    def _scheduler_job(self):
        if self.rodando:
            self._log("Scheduler: execução anterior ainda em andamento, pulando ciclo.")
            return
        # usa as mesmas datas do dia atual por padrão
        data_ini = date.today()
        data_fim = date.today()
        apenas_nao = True
        self._log("Scheduler: iniciando ciclo agendado...")
        def _run():
            try:
                executar_pipeline(self.cfg, data_ini, data_fim, apenas_nao, self._log)
            except Exception as e:
                self._log(f"❌ Erro no job agendado: {e}")
        threading.Thread(target=_run, daemon=True).start()

    def _start_scheduler(self):
        if not SCHED_OK or BackgroundScheduler is None:
            messagebox.showerror("Agendador ausente", f"APScheduler não instalado: {_SCHED_ERR}\nInstale com: pip install APScheduler")
            return False
        if self.scheduler:
            return True
        try:
            self.scheduler = BackgroundScheduler()
            interval = int(self.cfg.get("scheduler_interval_min", 60))
            self.scheduler.add_job(self._scheduler_job, 'interval', minutes=interval, id='ouvidoria_job', replace_existing=True)
            self.scheduler.start()
            self._log(f"✅ Agendador iniciado: a cada {interval} min")
            return True
        except Exception as e:
            self._log(f"❌ Falha ao iniciar agendador: {e}")
            self.scheduler = None
            return False

    def _stop_scheduler(self):
        if self.scheduler:
            try:
                self.scheduler.shutdown(wait=False)
            except Exception:
                pass
            self.scheduler = None
            self._log("🛑 Agendador parado.")

    def _toggle_scheduler(self, var: tk.BooleanVar, interval_var: tk.StringVar, btn: tk.Button):
        enabled = var.get()
        try:
            self.cfg["scheduler_interval_min"] = int(interval_var.get())
        except Exception:
            messagebox.showerror("Intervalo inválido", "Informe um número inteiro de minutos.")
            var.set(False)
            return
        self.cfg["scheduler_enabled"] = enabled
        salvar_config(self.cfg)
        if enabled:
            ok = self._start_scheduler()
            if ok:
                btn.configure(text="Parar Agendador")
            else:
                var.set(False)
        else:
            self._stop_scheduler()
            btn.configure(text="Iniciar Agendador")

    # ── Aba Planilha (visualização simples) ───────────────────────────────
    def _aba_planilha(self, nb):
        frm = tk.Frame(nb, bg=COR_CINZA)
        nb.add(frm, text="  📋 Planilha  ")

        top = tk.Frame(frm, bg=COR_CINZA)
        top.pack(fill="x", padx=12, pady=8)
        tk.Label(top, text="Filtro:", font=_fonte(10), bg=COR_CINZA).pack(side="left")
        self.planilha_filter_var = tk.StringVar()
        e = ttk.Entry(top, textvariable=self.planilha_filter_var, width=36)
        e.pack(side="left", padx=6)
        e.bind("<KeyRelease>", lambda ev: self._apply_planilha_filter())

        tk.Button(top, text="🔄 Atualizar", command=self._load_planilha, bg=COR_AZUL, fg=COR_BRANCO).pack(side="left", padx=8)

        # Treeview
        cols = COLUNAS_OUV
        tree_frm = tk.Frame(frm, bg=COR_CINZA)
        tree_frm.pack(fill="both", expand=True, padx=12, pady=(4,12))
        self.planilha_tree = ttk.Treeview(tree_frm, columns=cols, show='headings')
        for c in cols:
            self.planilha_tree.heading(c, text=c)
            self.planilha_tree.column(c, width=120, anchor='w')
        vsb = ttk.Scrollbar(tree_frm, orient='vertical', command=self.planilha_tree.yview)
        self.planilha_tree.configure(yscroll=vsb.set)
        vsb.pack(side='right', fill='y')
        self.planilha_tree.pack(fill='both', expand=True)

        self._planilha_rows = []
        self._load_planilha()

    def _load_planilha(self):
        pasta = self.cfg.get("pasta_base","ouvidorias")
        relat = os.path.join(pasta, "ouvidorias.xlsx")
        self._planilha_rows = []
        for i in self.planilha_tree.get_children():
            self.planilha_tree.delete(i)
        if not os.path.exists(relat):
            self._log("Nenhuma planilha encontrada para visualizar.")
            return
        try:
            wb = load_workbook(relat, read_only=True)
            ws = wb.active
            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            # map indices
            for row in ws.iter_rows(min_row=2, values_only=True):
                vals = [str(row[headers.index(c)]) if c in headers and row[headers.index(c)] is not None else "" for c in COLUNAS_OUV]
                self._planilha_rows.append(vals)
            wb.close()
            for r in self._planilha_rows:
                self.planilha_tree.insert('', 'end', values=r)
            self._log(f"Planilha carregada: {relat} ({len(self._planilha_rows)} linhas)")
        except Exception as e:
            self._log(f"Erro ao carregar planilha: {e}")

    def _apply_planilha_filter(self):
        q = self.planilha_filter_var.get().strip().lower()
        for i in self.planilha_tree.get_children():
            self.planilha_tree.delete(i)
        if not q:
            for r in self._planilha_rows:
                self.planilha_tree.insert('', 'end', values=r)
            return
        for r in self._planilha_rows:
            if any(q in (str(cell) or "").lower() for cell in r):
                self.planilha_tree.insert('', 'end', values=r)

# =============================================================================
# ENTRY POINT
# =============================================================================
if __name__ == "__main__":
    app = App()
    app.mainloop()
