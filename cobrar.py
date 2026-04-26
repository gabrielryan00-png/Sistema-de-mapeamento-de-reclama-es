import os
import json
import smtplib
import logging
from datetime import datetime, timedelta
from email.message import EmailMessage

try:
    from openpyxl import load_workbook
except ImportError:
    raise ImportError("openpyxl não está instalado. Execute: pip install openpyxl")

from banco import buscar_email_unidade
from constantes import CATALOGO_USF, normalizar, identificar_unidade

# Configurar logging persistente
LOG_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logs")
os.makedirs(LOG_DIR, exist_ok=True)
LOG_FILE = os.path.join(LOG_DIR, f"cobranca_{datetime.now().strftime('%Y-%m-%d')}.log")

# Configurar logger
logger = logging.getLogger("cobranca")
logger.setLevel(logging.INFO)

# Handler para arquivo
if not logger.handlers:
    fh = logging.FileHandler(LOG_FILE, encoding="utf-8")
    fh.setLevel(logging.INFO)
    formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
    fh.setFormatter(formatter)
    logger.addHandler(fh)

EMAIL = ""   # fallback — carregado de config.json em tempo de execução
SENHA = ""   # fallback — carregado de config.json em tempo de execução

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

def _creds():
    """Retorna (email, senha) lidos de config.json no momento da chamada."""
    try:
        with open(os.path.join(BASE_DIR, 'config.json'), encoding='utf-8') as f:
            cfg = json.load(f)
        return cfg.get('email', EMAIL), cfg.get('senha_app', SENHA)
    except Exception:
        return EMAIL, SENHA
PLANILHA = os.path.join(BASE_DIR, "ouvidorias", "ouvidorias.xlsx")

# Configurações de cobrança
DIAS_MINIMOS_COBRANCA = 3
DIAS_ENTRE_COBRACAS = 7  # Evita recobrança dentro de 7 dias


def enviar_email(destino, assunto, corpo, log_func=None):
    """Envia email com tratamento de erro."""
    if log_func is None:
        log_func = lambda msg: logger.info(msg)

    email, senha = _creds()
    try:
        msg = EmailMessage()
        msg["Subject"] = assunto
        msg["From"] = email
        msg["To"] = destino
        msg.set_content(corpo)
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(email, senha)
            smtp.send_message(msg)
        return True
    except Exception as e:
        log_func(f"❌ Erro ao enviar email para {destino}: {e}")
        logger.error(f"Erro ao enviar email para {destino}: {e}")
        return False


def executar_cobranca(log_func=None, unidades_filtro=None):
    """Executa cobranças de ouvidorias pendentes baseado na planilha.

    unidades_filtro: lista de nomes oficiais de unidade para restringir o envio.
                     None = todas as unidades elegíveis.
    """
    if log_func is None:
        log_func = logger.info

    stats = {
        "total_enviadas": 0,
        "indeterminadas_puladas": 0,
        "nao_pendentes_puladas": 0,
        "recobranca_pulada": 0,
        "unidade_nao_reconhecida": 0,
        "sem_email": 0,
        "erros_envio": 0,
    }

    if not os.path.exists(PLANILHA):
        msg = "❌ Planilha não encontrada."
        log_func(msg)
        logger.error(msg)
        return stats

    try:
        wb = load_workbook(PLANILHA)
        ws = wb["Ouvidorias"]
    except Exception as e:
        msg = f"❌ Erro ao abrir Excel: {e}"
        log_func(msg)
        logger.error(msg)
        return stats

    headers = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}

    required_cols = ["Protocolo", "Unidade", "Status", "Prazo Resposta", "Observações", "Data Última Cobrança"]
    missing = [col for col in required_cols if col not in headers]
    if missing:
        msg = f"❌ Colunas ausentes no Excel: {missing}"
        log_func(msg)
        logger.error(msg)
        return stats

    hoje = datetime.today().date()

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        try:
            col_proto          = headers["Protocolo"]
            col_unidade        = headers["Unidade"]
            col_status         = headers["Status"]
            col_prazo          = headers["Prazo Resposta"]
            col_observacoes    = headers["Observações"]
            col_ultima_cobranca = headers["Data Última Cobrança"]

            protocolo            = row[col_proto - 1].value
            unidade              = row[col_unidade - 1].value
            status               = row[col_status - 1].value
            prazo                = row[col_prazo - 1].value
            observacoes          = row[col_observacoes - 1].value or ""
            data_ultima_cobranca = row[col_ultima_cobranca - 1].value

            # Filtro 1: Protocolo válido
            if protocolo is None or "NÃO" in str(protocolo).upper():
                stats["nao_pendentes_puladas"] += 1
                continue

            # Filtro 2: Ignorar Indeterminadas
            if "Indeterminada" in str(observacoes):
                msg = f"⏭️  Protocolo {protocolo}: Pulando indeterminada"
                log_func(msg)
                logger.info(msg)
                stats["indeterminadas_puladas"] += 1
                continue

            # Filtro 3: Status deve ser PENDENTE ou ENCAMINHADA
            status_norm = normalizar(str(status))
            if status_norm not in ("pendente", "encaminhada"):
                stats["nao_pendentes_puladas"] += 1
                continue

            # Filtro 4: Validar unidade
            unidade_oficial = identificar_unidade(str(unidade))
            if not unidade_oficial:
                msg = f"❌ Unidade não reconhecida: {unidade} (Protocolo {protocolo})"
                log_func(msg)
                logger.warning(msg)
                stats["unidade_nao_reconhecida"] += 1
                continue

            # Filtro 5: Restringir às unidades selecionadas (se informado)
            if unidades_filtro and unidade_oficial not in unidades_filtro:
                stats["nao_pendentes_puladas"] += 1
                continue

            # Filtro 6: Verificar data do prazo
            if not prazo:
                msg = f"⚠️  Protocolo {protocolo}: Sem data de prazo"
                log_func(msg)
                logger.warning(msg)
                stats["nao_pendentes_puladas"] += 1
                continue

            try:
                if isinstance(prazo, datetime):
                    data_prazo = prazo.date()
                else:
                    data_prazo = datetime.strptime(str(prazo), "%d/%m/%Y").date()
            except Exception as e:
                msg = f"❌ Data inválida ({prazo}) - Protocolo {protocolo}: {e}"
                log_func(msg)
                logger.error(msg)
                stats["nao_pendentes_puladas"] += 1
                continue

            dias = (hoje - data_prazo).days

            # Filtro 7: Mínimo de dias vencido
            if dias < DIAS_MINIMOS_COBRANCA:
                stats["nao_pendentes_puladas"] += 1
                continue

            # Filtro 8: Evitar recobrança frequente
            if data_ultima_cobranca:
                try:
                    if isinstance(data_ultima_cobranca, datetime):
                        data_ult = data_ultima_cobranca.date()
                    else:
                        data_ult = datetime.strptime(str(data_ultima_cobranca), "%d/%m/%Y").date()
                    dias_desde_cobranca = (hoje - data_ult).days
                    if dias_desde_cobranca < DIAS_ENTRE_COBRACAS:
                        msg = f"⏭️  Protocolo {protocolo}: Recobrança pulada ({dias_desde_cobranca} dias desde última cobrança)"
                        log_func(msg)
                        logger.info(msg)
                        stats["recobranca_pulada"] += 1
                        continue
                except Exception:
                    pass

            # Buscar email da unidade
            email = buscar_email_unidade(unidade_oficial)
            if not email:
                msg = f"❌ Sem email para unidade: {unidade_oficial} (Protocolo {protocolo})"
                log_func(msg)
                logger.warning(msg)
                stats["sem_email"] += 1
                continue

            # Montar email de cobrança
            assunto = f"Cobrança Ouvidoria - Protocolo {protocolo} ({dias} dias vencido)"
            corpo = f"""Prezado(a),

A ouvidoria de protocolo {protocolo} está pendente há {dias} dias e não foi respondida.

Favor verificar e responder conforme protocolo estabelecido.

Protocolo: {protocolo}
Unidade: {unidade_oficial}
Prazo vencido: {dias} dias

Atenciosamente,
Sistema de Ouvidoria
"""

            # Enviar email
            if enviar_email(email, assunto, corpo, log_func):
                log_func(f"📨 Cobrado: {unidade_oficial} - {protocolo} ({email}) - {dias} dias vencido")
                logger.info(f"📨 Cobrado: {unidade_oficial} - {protocolo} ({email}) - {dias} dias vencido")
                stats["total_enviadas"] += 1

                # Atualizar "Data Última Cobrança" na planilha
                try:
                    row[col_ultima_cobranca - 1].value = hoje
                except Exception as e:
                    logger.error(f"Erro ao atualizar data de cobrança: {e}")
            else:
                log_func(f"❌ Falha ao enviar cobrança para {protocolo}")
                logger.error(f"❌ Falha ao enviar cobrança para {protocolo}")
                stats["erros_envio"] += 1

        except Exception as e:
            msg = f"❌ Erro ao processar linha {row_num}: {e}"
            log_func(msg)
            logger.error(msg)

    # Salvar planilha atualizada
    try:
        wb.save(PLANILHA)
        msg = "✅ Planilha salva com sucesso"
        log_func(msg)
        logger.info(msg)
    except Exception as e:
        msg = f"❌ Erro ao salvar Excel: {e}"
        log_func(msg)
        logger.error(msg)

    # Log de resumo
    resumo = f"\n{'='*50}\n✅ RESUMO DE COBRANÇAS\n"
    resumo += f"  Total enviadas: {stats['total_enviadas']}\n"
    resumo += f"  Indeterminadas puladas: {stats['indeterminadas_puladas']}\n"
    resumo += f"  Não pendentes puladas: {stats['nao_pendentes_puladas']}\n"
    resumo += f"  Recobrança pulada: {stats['recobranca_pulada']}\n"
    resumo += f"  Unidade não reconhecida: {stats['unidade_nao_reconhecida']}\n"
    resumo += f"  Sem email: {stats['sem_email']}\n"
    resumo += f"  Erros de envio: {stats['erros_envio']}\n"
    resumo += f"  Log: {LOG_FILE}\n"
    resumo += f"{'='*50}\n"
    log_func(resumo)
    logger.info(resumo)

    return stats


if __name__ == "__main__":
    logger.info(f"Ciclo de cobrança iniciado - {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    executar_cobranca()
    logger.info(f"Ciclo de cobrança finalizado - {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")

