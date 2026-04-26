"""Servidor FastAPI local para o Dashboard de Ouvidorias."""
import os
import json
import time
import uuid
import threading
from datetime import datetime, date

_BASE         = os.path.dirname(os.path.abspath(__file__))
DASHBOARD_DIR = os.path.join(_BASE, 'dashboard')
CONFIG_FILE   = os.path.join(_BASE, 'config.json')
PORT          = 7731

try:
    from fastapi import FastAPI, HTTPException, Request, BackgroundTasks
    from fastapi.responses import HTMLResponse, JSONResponse
    from fastapi.middleware.cors import CORSMiddleware
    import uvicorn
    FASTAPI_OK = True
except ImportError:
    FASTAPI_OK = False

if FASTAPI_OK:
    app = FastAPI(title="Ouvidoria Dashboard API", docs_url=None, redoc_url=None)

    app.add_middleware(
        CORSMiddleware,
        allow_origins=["*"],
        allow_methods=["GET", "POST", "PATCH", "DELETE", "OPTIONS"],
        allow_headers=["*"],
    )

    # ── helpers ────────────────────────────────────────────────────────────────
    def _cfg():
        try:
            with open(CONFIG_FILE, encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return {}

    def _excel():
        cfg = _cfg()
        pasta = cfg.get('pasta_base', 'ouvidorias')
        return os.path.join(_BASE, pasta, 'ouvidorias.xlsx')

    # ── jobs (email processing) ────────────────────────────────────────────────
    _JOBS: dict = {}
    _PROC_LOCK  = threading.Lock()

    def _run_processar_emails(job_id: str, data_ini_str: str, data_fim_str: str):
        log = _JOBS[job_id]['log']
        def lf(msg): log.append(str(msg))
        try:
            import ouvidoriagmail as _gm
            data_ini = datetime.strptime(data_ini_str, '%Y-%m-%d').date()
            data_fim = datetime.strptime(data_fim_str, '%Y-%m-%d').date()
            lf(f"▶ Processando e-mails de {data_ini} até {data_fim}…")
            with _PROC_LOCK:
                result = _gm.processar_emails_api(data_ini, data_fim, log_func=lf)
            _JOBS[job_id]['status'] = 'done'
            _JOBS[job_id]['result'] = result or {}
        except Exception as e:
            _JOBS[job_id]['status'] = 'error'
            _JOBS[job_id]['result'] = {'error': str(e)}
            lf(f"❌ {e}")

    # ── JSX bundle ─────────────────────────────────────────────────────────────
    _JSX_FILES = ['data.jsx', 'mapa-unidades.jsx', 'split-triagem.jsx', 'command-center.jsx']

    _HTML_TEMPLATE = '''<!DOCTYPE html>
<html lang="pt-BR">
<head>
  <meta charset="UTF-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1.0"/>
  <title>Ouvidoria · Dashboard</title>
  <style>
    *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}
    html, body, #root {{ width: 100%; height: 100%; overflow: hidden; background: oklch(0.18 0.006 260); }}
    body {{ font-family: "Inter", ui-sans-serif, system-ui, -apple-system, sans-serif; }}
  </style>
  <link rel="preconnect" href="https://fonts.googleapis.com"/>
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin/>
  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=IBM+Plex+Mono:wght@400;500;700&display=swap" rel="stylesheet"/>
</head>
<body>
  <div id="root"></div>
  <script crossorigin src="https://unpkg.com/react@18/umd/react.production.min.js"></script>
  <script crossorigin src="https://unpkg.com/react-dom@18/umd/react-dom.production.min.js"></script>
  <script src="https://unpkg.com/@babel/standalone/babel.min.js"></script>
  <script type="text/babel" data-presets="react">
{jsx_bundle}

function App() {{
  React.useEffect(() => {{
    if (typeof window._OUV_FETCH === 'function') window._OUV_FETCH();
    const id = setInterval(() => {{
      if (typeof window._OUV_FETCH === 'function') window._OUV_FETCH();
    }}, 60000);
    return () => clearInterval(id);
  }}, []);
  return (
    <div style={{{{ width:'100vw', height:'100vh', display:'flex', overflow:'hidden' }}}}>
      <CommandCenter standalone />
    </div>
  );
}}

const root = ReactDOM.createRoot(document.getElementById('root'));
root.render(<App />);
  </script>
</body>
</html>'''

    # ── routes: static ─────────────────────────────────────────────────────────
    @app.get('/', response_class=HTMLResponse)
    async def index():
        parts = []
        for fname in _JSX_FILES:
            fpath = os.path.join(DASHBOARD_DIR, fname)
            try:
                with open(fpath, encoding='utf-8') as f:
                    parts.append(f'// ── {fname} ──\n' + f.read())
            except Exception as e:
                parts.append(f'// MISSING: {fname}: {e}')
        bundle = '\n\n'.join(parts)
        return _HTML_TEMPLATE.format(jsx_bundle=bundle)

    # ── routes: ouvidorias ─────────────────────────────────────────────────────
    @app.get('/api/ouvidorias')
    async def api_list():
        path = _excel()
        if not os.path.exists(path):
            return JSONResponse([])
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True)
            sheet = 'Ouvidorias' if 'Ouvidorias' in wb.sheetnames else wb.sheetnames[0]
            ws = wb[sheet]
            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                item = {}
                for i, h in enumerate(headers):
                    if h:
                        v = row[i] if i < len(row) else None
                        item[h] = str(v) if v is not None else ''
                if any(v for v in item.values()):
                    rows.append(item)
            wb.close()
            return JSONResponse(rows)
        except Exception as e:
            raise HTTPException(500, str(e))

    @app.post('/api/ouvidorias', status_code=201)
    async def api_create(request: Request):
        data = await request.json() or {}
        if not data.get('Protocolo'):
            data['Protocolo'] = f"MANUAL-{int(time.time())}"
        ok = _append_row(data)
        return {'ok': ok, 'protocolo': data['Protocolo']}

    @app.patch('/api/ouvidorias/{proto:path}')
    async def api_update(proto: str, request: Request):
        updates = await request.json() or {}
        ok = _patch_row(proto, updates)
        if not ok:
            raise HTTPException(404, 'protocolo não encontrado')
        return {'ok': True}

    @app.delete('/api/ouvidorias/{proto:path}')
    async def api_delete(proto: str):
        ok = _delete_row(proto)
        if not ok:
            raise HTTPException(404, 'protocolo não encontrado')
        return {'ok': True}

    # ── routes: cobranças ──────────────────────────────────────────────────────
    @app.post('/api/cobrar')
    async def api_cobrar(request: Request):
        try:
            import cobrar as _cobrar
            body = await request.json() or {}
            unidades = body.get('unidades') or None
            logs = []
            stats = _cobrar.executar_cobranca(log_func=logs.append, unidades_filtro=unidades)
            return {'ok': True, 'stats': stats, 'log': logs}
        except Exception as e:
            raise HTTPException(500, str(e))

    # ── routes: processamento de e-mails (background job) ─────────────────────
    @app.post('/api/processar-emails', status_code=202)
    async def api_processar_emails(request: Request, background_tasks: BackgroundTasks):
        body = await request.json() or {}
        hoje = date.today().strftime('%Y-%m-%d')
        data_ini = body.get('data_inicial') or hoje
        data_fim = body.get('data_final')   or hoje
        job_id = uuid.uuid4().hex[:8]
        _JOBS[job_id] = {'status': 'running', 'log': [], 'result': None}
        background_tasks.add_task(_run_processar_emails, job_id, data_ini, data_fim)
        return {'job_id': job_id}

    @app.get('/api/jobs/{job_id}')
    async def api_job_status(job_id: str):
        job = _JOBS.get(job_id)
        if not job:
            raise HTTPException(404, 'job não encontrado')
        return job

    # ── catch-all static (deve ficar APÓS todas as rotas /api/*) ──────────────
    @app.get('/{fn:path}')
    async def static_file(fn: str):
        from fastapi.responses import FileResponse
        path = os.path.join(DASHBOARD_DIR, fn)
        if os.path.isfile(path):
            return FileResponse(path)
        raise HTTPException(404)

    # ── Excel helpers ──────────────────────────────────────────────────────────
    _EXTRA_COLS = ['Reclamante', 'Canal']

    def _ensure_cols(ws):
        from openpyxl.styles import Font, PatternFill
        cur = [ws.cell(1, i).value for i in range(1, ws.max_column + 2)]
        for col in _EXTRA_COLS:
            if col not in cur:
                idx = ws.max_column + 1
                c = ws.cell(1, idx, col)
                c.font  = Font(bold=True, color='FFFFFF', name='Arial', size=11)
                c.fill  = PatternFill('solid', start_color='2B5597')

    def _append_row(data: dict) -> bool:
        try:
            from openpyxl import load_workbook, Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            path = _excel()
            os.makedirs(os.path.dirname(path) if os.path.dirname(path) else '.', exist_ok=True)
            if os.path.exists(path):
                wb = load_workbook(path)
            else:
                wb = Workbook(); wb.remove(wb.active)
            if 'Ouvidorias' not in wb.sheetnames:
                from constantes import COLUNAS_OUV
                ws = wb.create_sheet('Ouvidorias')
                for i, col in enumerate(COLUNAS_OUV + _EXTRA_COLS, 1):
                    c = ws.cell(1, i, col)
                    c.font  = Font(bold=True, color='FFFFFF', name='Arial', size=11)
                    c.fill  = PatternFill('solid', start_color='2B5597')
                    c.alignment = Alignment(horizontal='center', vertical='center')
                ws.freeze_panes = 'A2'
            else:
                ws = wb['Ouvidorias']
                _ensure_cols(ws)
            headers = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]
            row_n   = ws.max_row + 1
            for col, val in data.items():
                if col in headers:
                    ws.cell(row_n, headers.index(col) + 1, str(val) if val else '')
            wb.save(path); wb.close()
            return True
        except Exception as e:
            print(f'[Dashboard] Erro ao criar linha: {e}')
            return False

    def _patch_row(proto: str, updates: dict) -> bool:
        try:
            from openpyxl import load_workbook
            path = _excel()
            if not os.path.exists(path):
                return False
            wb = load_workbook(path)
            if 'Ouvidorias' not in wb.sheetnames:
                wb.close(); return False
            ws = wb['Ouvidorias']
            _ensure_cols(ws)
            headers = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]
            pcol    = next((i + 1 for i, h in enumerate(headers) if h == 'Protocolo'), None)
            if not pcol:
                wb.close(); return False
            for row in range(2, ws.max_row + 1):
                if str(ws.cell(row, pcol).value or '').strip() == proto.strip():
                    headers = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]
                    for col, val in updates.items():
                        if col in headers:
                            ws.cell(row, headers.index(col) + 1, str(val) if val else '')
                    wb.save(path); wb.close()
                    return True
            wb.close(); return False
        except Exception as e:
            print(f'[Dashboard] Erro ao atualizar linha: {e}')
            return False

    def _delete_row(proto: str) -> bool:
        try:
            from openpyxl import load_workbook
            path = _excel()
            if not os.path.exists(path):
                return False
            wb = load_workbook(path)
            if 'Ouvidorias' not in wb.sheetnames:
                wb.close(); return False
            ws = wb['Ouvidorias']
            headers = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]
            pcol    = next((i + 1 for i, h in enumerate(headers) if h == 'Protocolo'), None)
            if not pcol:
                wb.close(); return False
            for row in range(2, ws.max_row + 1):
                if str(ws.cell(row, pcol).value or '').strip() == proto.strip():
                    ws.delete_rows(row)
                    wb.save(path); wb.close()
                    return True
            wb.close(); return False
        except Exception as e:
            print(f'[Dashboard] Erro ao deletar linha: {e}')
            return False

# ── public start ───────────────────────────────────────────────────────────────
_started = False

def start_server(port: int = PORT) -> int:
    global _started
    if not FASTAPI_OK:
        print('[Dashboard] FastAPI/uvicorn não instalado — execute: pip install fastapi uvicorn')
        return 0
    if _started:
        return port
    _started = True

    def _run():
        uvicorn.run(app, host='127.0.0.1', port=port, log_level='error')

    t = threading.Thread(target=_run, daemon=True)
    t.start()
    return port


if __name__ == '__main__':
    p = start_server()
    print(f'Dashboard em http://localhost:{p}')
    try:
        while True:
            time.sleep(3600)
    except KeyboardInterrupt:
        pass
